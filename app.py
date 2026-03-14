Sí. Acá va **el código Python completo**, listo para copiar y pegar en `app.py`.

```python
import streamlit as st
import pandas as pd
import altair as alt
from datetime import datetime, date
from supabase import create_client, Client
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =========================================================
# CONFIG
# =========================================================
st.set_page_config(
    page_title="Calculadora de Condición Física",
    page_icon="💪",
    layout="wide"
)

st.markdown("""
<style>
[data-testid="stAppViewContainer"] {
    background-color: #f7f8fa;
}

[data-testid="stHeader"] {
    background: rgba(0, 0, 0, 0);
}

[data-testid="stToolbar"] {
    right: 1rem;
}

.block-container {
    padding-top: 1.5rem;
    padding-bottom: 2rem;
    max-width: 1400px;
}

.result-card {
    padding: 14px 16px;
    border-radius: 10px;
    font-size: 18px;
    font-weight: 700;
    margin-top: 8px;
    margin-bottom: 10px;
}

.motivo-box {
    background-color: #ffffff;
    border: 1px solid #e6e9ef;
    border-radius: 10px;
    padding: 12px 14px;
    margin-bottom: 12px;
}

.reco-box {
    background-color: #eef7ef;
    border: 1px solid #d4ead7;
    border-radius: 10px;
    padding: 12px 14px;
    margin-bottom: 14px;
}
</style>
""", unsafe_allow_html=True)

# =========================================================
# SUPABASE
# =========================================================
SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# =========================================================
# TABLAS NORMATIVAS
# =========================================================
TABLA_CAMINATA_6M = {
    150: {
        40: {2.5: 436, 10: 470, 25: 511, 50: 555, 75: 592, 90: 631, 97.5: 679},
        50: {2.5: 434, 10: 468, 25: 509, 50: 553, 75: 590, 90: 629, 97.5: 677},
        60: {2.5: 414, 10: 448, 25: 489, 50: 533, 75: 570, 90: 609, 97.5: 656},
        70: {2.5: 364, 10: 397, 25: 439, 50: 483, 75: 520, 90: 558, 97.5: 606},
        80: {2.5: 313, 10: 347, 25: 388, 50: 432, 75: 469, 90: 508, 97.5: 556},
    },
    160: {
        40: {2.5: 455, 10: 489, 25: 530, 50: 574, 75: 611, 90: 650, 97.5: 697},
        50: {2.5: 453, 10: 487, 25: 528, 50: 572, 75: 609, 90: 648, 97.5: 695},
        60: {2.5: 433, 10: 466, 25: 508, 50: 552, 75: 588, 90: 627, 97.5: 675},
        70: {2.5: 382, 10: 416, 25: 457, 50: 501, 75: 538, 90: 577, 97.5: 625},
        80: {2.5: 332, 10: 366, 25: 407, 50: 451, 75: 488, 90: 526, 97.5: 574},
    },
    170: {
        40: {2.5: 474, 10: 507, 25: 549, 50: 593, 75: 629, 90: 668, 97.5: 716},
        50: {2.5: 472, 10: 505, 25: 546, 50: 590, 75: 627, 90: 666, 97.5: 714},
        60: {2.5: 451, 10: 485, 25: 526, 50: 570, 75: 607, 90: 646, 97.5: 694},
        70: {2.5: 401, 10: 435, 25: 476, 50: 520, 75: 557, 90: 595, 97.5: 643},
        80: {2.5: 351, 10: 384, 25: 425, 50: 469, 75: 506, 90: 545, 97.5: 593},
    },
    180: {
        40: {2.5: 492, 10: 526, 25: 567, 50: 611, 75: 648, 90: 687, 97.5: 735},
        50: {2.5: 490, 10: 524, 25: 565, 50: 609, 75: 646, 90: 685, 97.5: 733},
        60: {2.5: 470, 10: 503, 25: 545, 50: 589, 75: 626, 90: 664, 97.5: 712},
        70: {2.5: 419, 10: 453, 25: 494, 50: 538, 75: 575, 90: 614, 97.5: 662},
        80: {2.5: 369, 10: 403, 25: 444, 50: 488, 75: 525, 90: 564, 97.5: 611},
    },
    190: {
        40: {2.5: 511, 10: 544, 25: 586, 50: 630, 75: 667, 90: 705, 97.5: 753},
        50: {2.5: 509, 10: 542, 25: 584, 50: 628, 75: 665, 90: 703, 97.5: 751},
        60: {2.5: 488, 10: 522, 25: 563, 50: 607, 75: 644, 90: 683, 97.5: 731},
        70: {2.5: 438, 10: 472, 25: 513, 50: 557, 75: 594, 90: 633, 97.5: 680},
        80: {2.5: 388, 10: 421, 25: 463, 50: 507, 75: 544, 90: 582, 97.5: 630},
    }
}

TABLA_PRENSION = {
    "Hombre": {
        "20-24": {5: 33.9, 10: 36.8, 20: 40.5, 30: 43.2, 40: 45.7, 50: 48.0, 60: 50.4, 70: 52.9, 80: 56.0, 90: 60.1, 95: 63.6},
        "25-29": {5: 35.5, 10: 38.5, 20: 42.1, 30: 44.8, 40: 47.1, 50: 49.3, 60: 51.5, 70: 53.9, 80: 56.7, 90: 60.7, 95: 64.0},
        "30-34": {5: 35.0, 10: 38.3, 20: 42.2, 30: 45.0, 40: 47.4, 50: 49.7, 60: 52.0, 70: 54.4, 80: 57.4, 90: 61.5, 95: 64.9},
        "35-39": {5: 33.8, 10: 37.3, 20: 41.5, 30: 44.5, 40: 47.1, 50: 49.5, 60: 51.9, 70: 54.4, 80: 57.5, 90: 61.8, 95: 65.3},
        "40-44": {5: 32.3, 10: 36.0, 20: 40.4, 30: 43.6, 40: 46.3, 50: 48.8, 60: 51.2, 70: 53.9, 80: 57.1, 90: 61.5, 95: 65.1},
        "45-49": {5: 30.6, 10: 34.4, 20: 39.0, 30: 42.3, 40: 45.1, 50: 47.6, 60: 50.2, 70: 52.9, 80: 56.2, 90: 60.7, 95: 64.4},
        "50-54": {5: 28.9, 10: 32.8, 20: 37.4, 30: 40.7, 40: 43.5, 50: 46.2, 60: 48.8, 70: 51.6, 80: 54.8, 90: 59.4, 95: 63.1},
        "55-59": {5: 27.2, 10: 31.0, 20: 35.6, 30: 38.9, 40: 41.7, 50: 44.4, 60: 47.0, 70: 49.8, 80: 53.1, 90: 57.7, 95: 61.4},
        "60-64": {5: 25.5, 10: 29.1, 20: 33.6, 30: 36.9, 40: 39.7, 50: 42.4, 60: 45.0, 70: 47.8, 80: 51.1, 90: 55.6, 95: 59.3},
        "65-69": {5: 23.7, 10: 27.2, 20: 31.5, 30: 34.7, 40: 37.5, 50: 40.1, 60: 42.8, 70: 45.6, 80: 48.8, 90: 53.2, 95: 56.8},
        "70-74": {5: 21.9, 10: 25.2, 20: 29.3, 30: 32.4, 40: 35.1, 50: 37.7, 60: 40.3, 70: 43.1, 80: 46.3, 90: 50.6, 95: 54.1},
        "75-79": {5: 20.0, 10: 23.1, 20: 27.0, 30: 29.9, 40: 32.5, 50: 35.1, 60: 37.6, 70: 40.3, 80: 43.5, 90: 47.7, 95: 51.1},
        "80-84": {5: 18.0, 10: 20.8, 20: 24.5, 30: 27.3, 40: 29.8, 50: 32.3, 60: 34.8, 70: 37.5, 80: 40.5, 90: 44.7, 95: 48.0},
        "85-89": {5: 15.9, 10: 18.5, 20: 21.9, 30: 24.6, 40: 27.0, 50: 29.4, 60: 31.8, 70: 34.4, 80: 37.4, 90: 41.5, 95: 44.6},
        "90-94": {5: 13.7, 10: 16.1, 20: 19.2, 30: 21.7, 40: 24.0, 50: 26.3, 60: 28.7, 70: 31.2, 80: 34.2, 90: 38.1, 95: 41.2},
        "95-99": {5: 11.3, 10: 13.5, 20: 16.4, 30: 18.8, 40: 20.9, 50: 23.1, 60: 25.4, 70: 27.9, 80: 30.8, 90: 34.6, 95: 37.5},
        "+100": {5: 8.8, 10: 10.8, 20: 13.5, 30: 15.7, 40: 17.8, 50: 19.8, 60: 22.0, 70: 24.5, 80: 27.2, 90: 30.9, 95: 33.8},
    },
    "Mujer": {
        "20-24": {5: 19.7, 10: 21.7, 20: 24.0, 30: 25.7, 40: 27.2, 50: 28.6, 60: 30.0, 70: 31.6, 80: 33.6, 90: 36.6, 95: 39.1},
        "25-29": {5: 20.0, 10: 22.0, 20: 24.5, 30: 26.3, 40: 27.9, 50: 29.4, 60: 30.9, 70: 32.6, 80: 34.6, 90: 37.4, 95: 39.7},
        "30-34": {5: 19.6, 10: 21.8, 20: 24.4, 30: 26.4, 40: 28.1, 50: 29.7, 60: 31.3, 70: 33.1, 80: 35.2, 90: 38.0, 95: 40.4},
        "35-39": {5: 19.0, 10: 21.3, 20: 24.1, 30: 26.2, 40: 28.0, 50: 29.7, 60: 31.4, 70: 33.2, 80: 35.4, 90: 38.4, 95: 40.8},
        "40-44": {5: 18.3, 10: 20.7, 20: 23.7, 30: 25.8, 40: 27.6, 50: 29.4, 60: 31.1, 70: 33.0, 80: 35.2, 90: 38.3, 95: 40.8},
        "45-49": {5: 17.6, 10: 20.1, 20: 23.1, 30: 25.2, 40: 27.1, 50: 28.9, 60: 30.6, 70: 32.5, 80: 34.8, 90: 37.9, 95: 40.4},
        "50-54": {5: 16.9, 10: 19.4, 20: 22.4, 30: 24.5, 40: 26.4, 50: 28.2, 60: 29.9, 70: 31.8, 80: 34.0, 90: 37.1, 95: 39.7},
        "55-59": {5: 16.1, 10: 18.5, 20: 21.5, 30: 23.7, 40: 25.5, 50: 27.3, 60: 29.0, 70: 30.9, 80: 33.0, 90: 36.1, 95: 38.6},
        "60-64": {5: 15.2, 10: 17.6, 20: 20.6, 30: 22.7, 40: 24.5, 50: 26.2, 60: 27.9, 70: 29.7, 80: 31.8, 90: 34.9, 95: 37.4},
        "65-69": {5: 14.3, 10: 16.6, 20: 19.5, 30: 21.6, 40: 23.3, 50: 25.0, 60: 26.6, 70: 28.4, 80: 30.5, 90: 33.4, 95: 35.8},
        "70-74": {5: 13.2, 10: 15.5, 20: 18.3, 30: 20.3, 40: 22.0, 50: 23.6, 60: 25.2, 70: 26.9, 80: 28.9, 90: 31.8, 95: 34.1},
        "75-79": {5: 12.0, 10: 14.3, 20: 17.0, 30: 18.9, 40: 20.5, 50: 22.1, 60: 23.6, 70: 25.2, 80: 27.2, 90: 29.9, 95: 32.2},
        "80-84": {5: 10.7, 10: 12.9, 20: 15.5, 30: 17.4, 40: 18.9, 50: 20.4, 60: 21.9, 70: 23.5, 80: 25.3, 90: 28.0, 95: 30.2},
        "85-89": {5: 9.3, 10: 11.4, 20: 13.9, 30: 15.7, 40: 17.2, 50: 18.6, 60: 20.0, 70: 21.5, 80: 23.3, 90: 25.9, 95: 28.0},
        "90-94": {5: 7.8, 10: 9.8, 20: 12.2, 30: 13.9, 40: 15.3, 50: 16.7, 60: 18.0, 70: 19.5, 80: 21.2, 90: 23.6, 95: 25.7},
        "95-99": {5: 6.1, 10: 8.0, 20: 10.3, 30: 11.9, 40: 13.3, 50: 14.6, 60: 15.9, 70: 17.3, 80: 18.9, 90: 21.2, 95: 23.2},
        "+100": {5: 4.2, 10: 6.1, 20: 8.3, 30: 9.8, 40: 11.2, 50: 12.4, 60: 13.6, 70: 14.9, 80: 16.5, 90: 18.7, 95: 20.6},
    }
}

TABLA_SILLA = {
    "Hombre": {
        "65-69": {10: 12, 20: 13, 30: 14, 40: 15, 50: 16, 60: 16, 70: 17, 80: 19, 90: 21, 100: 28},
        "70-74": {10: 11, 20: 13, 30: 14, 40: 15, 50: 15, 60: 16, 70: 17, 80: 18, 90: 20, 100: 29},
        "75-79": {10: 10, 20: 12, 30: 13, 40: 14, 50: 14, 60: 15, 70: 16, 80: 17, 90: 19, 100: 25},
        "80-84": {10: 9, 20: 10, 30: 11, 40: 12, 50: 14, 60: 15, 70: 16, 80: 17, 90: 18, 100: 22},
        "+84": {10: 9, 20: 9, 30: 12, 40: 13, 50: 14, 60: 14, 70: 16, 80: 18, 90: 20, 100: 21},
    },
    "Mujer": {
        "65-69": {10: 11, 20: 12, 30: 13, 40: 14, 50: 15, 60: 15, 70: 16, 80: 17, 90: 19, 100: 30},
        "70-74": {10: 10, 20: 12, 30: 12, 40: 13, 50: 14, 60: 15, 70: 16, 80: 17, 90: 19, 100: 27},
        "75-79": {10: 10, 20: 11, 30: 12, 40: 13, 50: 14, 60: 14, 70: 15, 80: 16, 90: 18, 100: 24},
        "80-84": {10: 9, 20: 10, 30: 11, 40: 12, 50: 13, 60: 14, 70: 15, 80: 16, 90: 18, 100: 24},
        "+84": {10: 6, 20: 8, 30: 9, 40: 11, 50: 12, 60: 14, 70: 14, 80: 16, 90: 17, 100: 18},
    }
}

# =========================================================
# BASE DE DATOS
# =========================================================
def guardar_evaluacion(paciente_nombre, sexo, edad, prueba, valor_medido, percentil, clasificacion):
    payload = {
        "fecha": datetime.now().strftime("%Y-%m-%d"),
        "paciente": str(paciente_nombre).strip(),
        "sexo": str(sexo).strip().lower(),
        "edad": int(edad),
        "prueba": str(prueba).strip(),
        "valor_medido": float(valor_medido),
        "percentil": round(float(percentil), 1) if percentil is not None else None,
        "clasificacion": str(clasificacion).strip()
    }
    return supabase.table("evaluaciones").insert(payload).execute()


def guardar_paciente(nombre, sexo, talla_m):
    nombre_limpio = str(nombre).strip()
    sexo_limpio = str(sexo).strip().lower()

    if not nombre_limpio:
        raise ValueError("El nombre del paciente está vacío.")

    if talla_m is None or float(talla_m) <= 0:
        raise ValueError("La talla debe ser mayor a 0.")

    respuesta = supabase.table("pacientes").select("id,nombre").execute()
    existentes = respuesta.data if respuesta.data else []

    for p in existentes:
        if str(p["nombre"]).strip().lower() == nombre_limpio.lower():
            raise ValueError("Ese paciente ya existe.")

    payload = {
        "nombre": nombre_limpio,
        "sexo": sexo_limpio,
        "talla_m": round(float(talla_m), 2)
    }
    return supabase.table("pacientes").insert(payload).execute()


def guardar_peso(paciente_id, fecha_medicion, peso_kg, talla_m):
    if paciente_id is None:
        raise ValueError("No se encontró el id del paciente.")

    if talla_m is None or float(talla_m) <= 0:
        raise ValueError("El paciente no tiene una talla válida cargada.")

    if peso_kg is None or float(peso_kg) <= 0:
        raise ValueError("El peso debe ser mayor a 0.")

    imc = round(float(peso_kg) / (float(talla_m) ** 2), 2)

    payload = {
        "paciente_id": int(paciente_id),
        "fecha": str(fecha_medicion),
        "peso_kg": float(peso_kg),
        "imc": imc
    }

    return supabase.table("seguimiento_peso").insert(payload).execute()


def clasificar_imc(imc_calculado):
    if imc_calculado < 18.5:
        return "Bajo peso", "🔵"
    elif imc_calculado < 25:
        return "Normal", "🟢"
    elif imc_calculado < 30:
        return "Sobrepeso", "🟡"
    else:
        return "Obesidad", "🔴"


def eliminar_evaluacion(id_registro):
    return supabase.table("evaluaciones").delete().eq("id", id_registro).execute()


def obtener_historial_paciente(paciente):
    try:
        respuesta = (
            supabase.table("evaluaciones")
            .select("*")
            .eq("paciente", str(paciente).strip())
            .order("fecha")
            .execute()
        )
        if respuesta.data:
            return pd.DataFrame(respuesta.data)
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Error al leer historial: {e}")
        return pd.DataFrame()


def obtener_historial_peso(paciente_id):
    try:
        respuesta = (
            supabase.table("seguimiento_peso")
            .select("id,paciente_id,fecha,peso_kg,imc,created_at")
            .eq("paciente_id", int(paciente_id))
            .order("fecha")
            .execute()
        )

        if respuesta.data:
            return pd.DataFrame(respuesta.data)

        return pd.DataFrame(columns=["fecha", "peso_kg", "imc"])
    except Exception as e:
        st.error(f"Error al leer historial de peso: {e}")
        return pd.DataFrame(columns=["fecha", "peso_kg", "imc"])


def obtener_pacientes():
    try:
        respuesta = (
            supabase.table("pacientes")
            .select("id,nombre,sexo,talla_m")
            .order("nombre")
            .execute()
        )
        return respuesta.data if respuesta.data else []
    except Exception as e:
        st.error(f"Error al leer pacientes: {e}")
        return []


def guardar_inbody(
    paciente_id,
    fecha_estudio,
    peso_kg,
    talla_m,
    grasa_corporal_pct,
    masa_muscular_kg,
    agua_corporal_pct,
    grasa_visceral,
    metabolismo_basal,
    observaciones
):
    imc_calculado = round(float(peso_kg) / (float(talla_m) ** 2), 2) if peso_kg and talla_m else None

    payload = {
        "paciente_id": int(paciente_id),
        "fecha": str(fecha_estudio),
        "peso_kg": float(peso_kg) if peso_kg is not None else None,
        "imc": float(imc_calculado) if imc_calculado is not None else None,
        "grasa_corporal_pct": float(grasa_corporal_pct) if grasa_corporal_pct is not None else None,
        "masa_muscular_kg": float(masa_muscular_kg) if masa_muscular_kg is not None else None,
        "agua_corporal_pct": float(agua_corporal_pct) if agua_corporal_pct is not None else None,
        "grasa_visceral": float(grasa_visceral) if grasa_visceral is not None else None,
        "metabolismo_basal": float(metabolismo_basal) if metabolismo_basal is not None else None,
        "observaciones": str(observaciones).strip() if observaciones else None
    }

    return supabase.table("inbody_registros").insert(payload).execute()


def obtener_historial_inbody(paciente_id):
    try:
        respuesta = (
            supabase.table("inbody_registros")
            .select("*")
            .eq("paciente_id", int(paciente_id))
            .order("fecha", desc=True)
            .execute()
        )

        if respuesta.data:
            return pd.DataFrame(respuesta.data)

        return pd.DataFrame()

    except Exception as e:
        st.error(f"Error al leer historial de composición corporal: {e}")
        return pd.DataFrame()


def obtener_ficha_paciente(paciente_nombre):
    try:
        paciente_nombre = str(paciente_nombre).strip()

        respuesta_paciente = (
            supabase.table("pacientes")
            .select("id,nombre,sexo,talla_m,created_at")
            .eq("nombre", paciente_nombre)
            .limit(1)
            .execute()
        )

        datos_paciente = respuesta_paciente.data[0] if respuesta_paciente.data else {}
        df_historial = obtener_historial_paciente(paciente_nombre)

        cantidad_evaluaciones = 0
        ultima_fecha = "-"
        ultima_clasificacion = "-"
        ultima_prueba = "-"

        if not df_historial.empty:
            if "fecha" in df_historial.columns:
                df_historial["fecha"] = pd.to_datetime(df_historial["fecha"], errors="coerce")
                df_historial = df_historial.sort_values("fecha", ascending=False)

            cantidad_evaluaciones = len(df_historial)

            if "fecha" in df_historial.columns and pd.notna(df_historial.iloc[0]["fecha"]):
                ultima_fecha = df_historial.iloc[0]["fecha"].strftime("%d-%m-%Y")

            if "clasificacion" in df_historial.columns and pd.notna(df_historial.iloc[0]["clasificacion"]):
                ultima_clasificacion = str(df_historial.iloc[0]["clasificacion"])

            if "prueba" in df_historial.columns and pd.notna(df_historial.iloc[0]["prueba"]):
                ultima_prueba = str(df_historial.iloc[0]["prueba"])

        return {
            "id": datos_paciente.get("id"),
            "nombre": datos_paciente.get("nombre", paciente_nombre),
            "sexo": datos_paciente.get("sexo", "-"),
            "talla_m": datos_paciente.get("talla_m"),
            "cantidad_evaluaciones": cantidad_evaluaciones,
            "ultima_fecha": ultima_fecha,
            "ultima_clasificacion": ultima_clasificacion,
            "ultima_prueba": ultima_prueba
        }

    except Exception as e:
        st.error(f"Error al cargar ficha del paciente: {e}")
        return {
            "id": None,
            "nombre": paciente_nombre,
            "sexo": "-",
            "talla_m": None,
            "cantidad_evaluaciones": 0,
            "ultima_fecha": "-",
            "ultima_clasificacion": "-",
            "ultima_prueba": "-"
        }


def preparar_df_exportacion(df):
    if df is None or df.empty:
        return pd.DataFrame()

    df_out = df.copy()

    for col in df_out.columns:
        if "fecha" in col.lower() or "created" in col.lower():
            try:
                df_out[col] = pd.to_datetime(df_out[col], errors="coerce")
                df_out[col] = df_out[col].dt.strftime("%Y-%m-%d")
            except Exception:
                pass

    return df_out


# =========================================================
# UTILIDADES CLÍNICAS - COMPOSICIÓN CORPORAL
# =========================================================
def clasificacion_grasa_corporal(sexo, grasa_pct):
    sexo = str(sexo).strip().lower()

    if grasa_pct is None or pd.isna(grasa_pct):
        return "Sin clasificar"

    if sexo == "hombre":
        if grasa_pct < 8:
            return "Muy bajo"
        elif grasa_pct < 19:
            return "Normal"
        elif grasa_pct <= 25:
            return "Alto"
        else:
            return "Obesidad"

    if sexo == "mujer":
        if grasa_pct < 21:
            return "Muy bajo"
        elif grasa_pct < 33:
            return "Normal"
        elif grasa_pct <= 39:
            return "Alto"
        else:
            return "Obesidad"

    return "Sin clasificar"


def clasificacion_agua_corporal(sexo, agua_pct):
    sexo = str(sexo).strip().lower()

    if agua_pct is None or pd.isna(agua_pct):
        return "Sin clasificar"

    if sexo == "hombre":
        if agua_pct < 50:
            return "Bajo"
        elif agua_pct <= 65:
            return "Normal"
        else:
            return "Alto"

    if sexo == "mujer":
        if agua_pct < 45:
            return "Bajo"
        elif agua_pct <= 60:
            return "Normal"
        else:
            return "Alto"

    return "Sin clasificar"


def clasificacion_grasa_visceral(grasa_visceral):
    if grasa_visceral is None or pd.isna(grasa_visceral):
        return "Sin clasificar"

    if grasa_visceral <= 9:
        return "Normal"
    elif grasa_visceral <= 14:
        return "Alto"
    else:
        return "Muy alto"


def calcular_masa_muscular_relativa_pct(peso_kg, masa_muscular_kg):
    if peso_kg is None or masa_muscular_kg is None:
        return None
    if pd.isna(peso_kg) or pd.isna(masa_muscular_kg):
        return None
    if float(peso_kg) <= 0:
        return None
    return round((float(masa_muscular_kg) / float(peso_kg)) * 100, 2)


def clasificacion_masa_muscular_relativa(sexo, musculo_relativo_pct):
    sexo = str(sexo).strip().lower()

    if musculo_relativo_pct is None or pd.isna(musculo_relativo_pct):
        return "Sin clasificar"

    if sexo == "hombre":
        if musculo_relativo_pct < 33:
            return "Bajo"
        elif musculo_relativo_pct <= 39:
            return "Normal"
        else:
            return "Alto"

    if sexo == "mujer":
        if musculo_relativo_pct < 24:
            return "Bajo"
        elif musculo_relativo_pct <= 30:
            return "Normal"
        else:
            return "Alto"

    return "Sin clasificar"


def color_estado_corporal(estado):
    mapa = {
        "Normal": ("#2e7d32", "#ffffff"),
        "Bajo peso": ("#1976d2", "#ffffff"),
        "Riesgo sarcopénico": ("#ef6c00", "#ffffff"),
        "Sobrepeso": ("#f9a825", "#1f1f1f"),
        "Sobrepeso muscular": ("#00897b", "#ffffff"),
        "Obesidad": ("#c62828", "#ffffff"),
        "Riesgo cardiometabólico": ("#ad1457", "#ffffff"),
        "Riesgo cardiometabólico moderado": ("#6a1b9a", "#ffffff"),
        "Sin clasificar": ("#757575", "#ffffff")
    }
    return mapa.get(estado, ("#757575", "#ffffff"))


def generar_recomendacion_corporal(estado, clasif_grasa, clasif_visceral, clasif_musculo):
    if estado in ["Obesidad", "Riesgo cardiometabólico", "Riesgo cardiometabólico moderado"]:
        return "Programa de reducción de grasa + ejercicio de fuerza + control cardiometabólico."
    if estado == "Riesgo sarcopénico":
        return "Priorizar ejercicio de fuerza, aumento de masa muscular y seguimiento funcional."
    if estado == "Sobrepeso muscular":
        return "Mantener masa muscular, controlar evolución y ajustar plan nutricional según objetivo."
    if estado == "Sobrepeso":
        return "Plan de control de peso con actividad física regular y seguimiento de composición corporal."
    if estado == "Bajo peso":
        return "Evaluar aporte nutricional y preservar o mejorar masa muscular."
    if estado == "Normal":
        if clasif_musculo == "Bajo":
            return "Estado general aceptable, pero conviene reforzar trabajo de fuerza."
        return "Mantener hábitos actuales y seguimiento periódico."
    return "Completar datos clínicos y repetir control."


def evaluar_perfil_morfofuncional(sexo, peso_kg, talla_m, grasa_pct, masa_muscular_kg, agua_pct, grasa_visceral):
    imc = round(float(peso_kg) / (float(talla_m) ** 2), 2) if peso_kg and talla_m else None
    clasif_imc = clasificar_imc(imc)[0] if imc is not None else "Sin clasificar"
    clasif_grasa = clasificacion_grasa_corporal(sexo, grasa_pct)
    clasif_agua = clasificacion_agua_corporal(sexo, agua_pct)
    clasif_visceral = clasificacion_grasa_visceral(grasa_visceral)
    musculo_rel_pct = calcular_masa_muscular_relativa_pct(peso_kg, masa_muscular_kg)
    clasif_musculo = clasificacion_masa_muscular_relativa(sexo, musculo_rel_pct)

    motivos = []

    if clasif_imc == "Bajo peso":
        motivos.append("IMC en rango de bajo peso")
    elif clasif_imc == "Sobrepeso":
        motivos.append("IMC en rango de sobrepeso")
    elif clasif_imc == "Obesidad":
        motivos.append("IMC en rango de obesidad")

    if clasif_grasa in ["Alto", "Obesidad"]:
        motivos.append("% grasa corporal elevada")

    if clasif_visceral in ["Alto", "Muy alto"]:
        motivos.append(f"grasa visceral {clasif_visceral.lower()}")

    if clasif_musculo == "Bajo":
        motivos.append("masa muscular relativa baja")
    elif clasif_musculo == "Normal":
        motivos.append("masa muscular relativa normal")
    elif clasif_musculo == "Alto":
        motivos.append("masa muscular relativa alta")

    estado = "Normal"

    if clasif_visceral == "Muy alto":
        estado = "Riesgo cardiometabólico"
    elif clasif_visceral == "Alto" and clasif_grasa in ["Alto", "Obesidad"]:
        estado = "Riesgo cardiometabólico moderado"
    elif clasif_imc == "Obesidad" or (clasif_imc in ["Sobrepeso", "Obesidad"] and clasif_grasa == "Obesidad"):
        estado = "Obesidad"
    elif clasif_imc == "Sobrepeso" and clasif_grasa in ["Alto", "Obesidad"] and clasif_musculo != "Alto":
        estado = "Sobrepeso"
    elif clasif_imc == "Sobrepeso" and clasif_musculo == "Alto" and clasif_grasa == "Normal":
        estado = "Sobrepeso muscular"
    elif clasif_imc == "Normal" and clasif_musculo == "Bajo":
        estado = "Riesgo sarcopénico"
    elif clasif_imc == "Bajo peso":
        estado = "Bajo peso"
    else:
        estado = "Normal"

    recomendacion = generar_recomendacion_corporal(
        estado=estado,
        clasif_grasa=clasif_grasa,
        clasif_visceral=clasif_visceral,
        clasif_musculo=clasif_musculo
    )

    return {
        "imc": imc,
        "clasif_imc": clasif_imc,
        "clasif_grasa": clasif_grasa,
        "clasif_agua": clasif_agua,
        "clasif_visceral": clasif_visceral,
        "musculo_rel_pct": musculo_rel_pct,
        "clasif_musculo": clasif_musculo,
        "estado": estado,
        "motivos": motivos,
        "recomendacion": recomendacion
    }


def enriquecer_historial_corporal(df, sexo, talla_m):
    if df is None or df.empty:
        return pd.DataFrame()

    df = df.copy()

    for col in [
        "peso_kg",
        "imc",
        "grasa_corporal_pct",
        "masa_muscular_kg",
        "agua_corporal_pct",
        "grasa_visceral",
        "metabolismo_basal"
    ]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    diagnosticos = []
    sugerencias = []
    motivos_lista = []
    clasif_imc_lista = []
    clasif_grasa_lista = []
    clasif_agua_lista = []
    clasif_visceral_lista = []
    musculo_rel_lista = []
    clasif_musculo_lista = []

    for _, row in df.iterrows():
        res = evaluar_perfil_morfofuncional(
            sexo=sexo,
            peso_kg=row.get("peso_kg"),
            talla_m=talla_m,
            grasa_pct=row.get("grasa_corporal_pct"),
            masa_muscular_kg=row.get("masa_muscular_kg"),
            agua_pct=row.get("agua_corporal_pct"),
            grasa_visceral=row.get("grasa_visceral")
        )

        diagnosticos.append(res["estado"])
        sugerencias.append(res["recomendacion"])
        motivos_lista.append(" | ".join(res["motivos"]) if res["motivos"] else "")
        clasif_imc_lista.append(res["clasif_imc"])
        clasif_grasa_lista.append(res["clasif_grasa"])
        clasif_agua_lista.append(res["clasif_agua"])
        clasif_visceral_lista.append(res["clasif_visceral"])
        musculo_rel_lista.append(res["musculo_rel_pct"])
        clasif_musculo_lista.append(res["clasif_musculo"])

    df["diagnostico_corporal"] = diagnosticos
    df["sugerencia_corporal"] = sugerencias
    df["motivos_corporal"] = motivos_lista
    df["clasif_imc"] = clasif_imc_lista
    df["clasif_grasa"] = clasif_grasa_lista
    df["clasif_agua"] = clasif_agua_lista
    df["clasif_visceral"] = clasif_visceral_lista
    df["musculo_rel_pct"] = musculo_rel_lista
    df["clasif_musculo"] = clasif_musculo_lista

    return df


def agregar_identificacion_paciente(df, ficha, origen=""):
    if df is None or df.empty:
        return pd.DataFrame(columns=[
            "Paciente",
            "PacienteID_Ficha",
            "Sexo",
            "Talla_m",
            "Origen"
        ])

    df_out = df.copy()

    rename_map = {}
    if "id" in df_out.columns:
        rename_map["id"] = "RegistroID"
    if "paciente_id" in df_out.columns:
        rename_map["paciente_id"] = "PacienteID_Registro"
    df_out = df_out.rename(columns=rename_map)

    columnas_identificacion = {
        "Paciente": ficha.get("nombre"),
        "PacienteID_Ficha": ficha.get("id"),
        "Sexo": ficha.get("sexo"),
        "Talla_m": ficha.get("talla_m"),
        "Origen": origen
    }

    for nombre_col, valor in reversed(list(columnas_identificacion.items())):
        if nombre_col in df_out.columns:
            df_out.drop(columns=[nombre_col], inplace=True)
        df_out.insert(0, nombre_col, valor)

    return df_out


def generar_df_analisis_cientifico(ficha, df_peso, df_inbody, df_eval):
    filas = []

    nombre = ficha.get("nombre")
    sexo = ficha.get("sexo")
    talla_m = ficha.get("talla_m")
    paciente_id = ficha.get("id")

    if df_peso is not None and not df_peso.empty:
        df_p = df_peso.copy()
        if "fecha" in df_p.columns:
            df_p["fecha"] = pd.to_datetime(df_p["fecha"], errors="coerce")

        for _, row in df_p.iterrows():
            filas.append({
                "PacienteID": paciente_id,
                "Paciente": nombre,
                "Sexo": sexo,
                "Talla_m": talla_m,
                "Fecha": row.get("fecha"),
                "TipoRegistro": "Peso_IMC",
                "Prueba": None,
                "ValorMedido": None,
                "Percentil": None,
                "ClasificacionFuncional": None,
                "Peso_kg": row.get("peso_kg"),
                "IMC": row.get("imc"),
                "GrasaCorporal_pct": None,
                "MasaMuscular_kg": None,
                "AguaCorporal_pct": None,
                "GrasaVisceral": None,
                "MetabolismoBasal": None,
                "MusculoRelativo_pct": None,
                "Clasif_IMC": clasificar_imc(row.get("imc"))[0] if pd.notna(row.get("imc")) else None,
                "Clasif_Grasa": None,
                "Clasif_Agua": None,
                "Clasif_Visceral": None,
                "Clasif_Musculo": None,
                "DiagnosticoCorporal": None,
                "SugerenciaCorporal": None,
                "MotivosCorporal": None,
                "Observaciones": None
            })

    if df_inbody is not None and not df_inbody.empty:
        df_c = enriquecer_historial_corporal(df_inbody, str(sexo).strip().lower(), talla_m)
        if "fecha" in df_c.columns:
            df_c["fecha"] = pd.to_datetime(df_c["fecha"], errors="coerce")

        for _, row in df_c.iterrows():
            filas.append({
                "PacienteID": paciente_id,
                "Paciente": nombre,
                "Sexo": sexo,
                "Talla_m": talla_m,
                "Fecha": row.get("fecha"),
                "TipoRegistro": "Composicion_Corporal",
                "Prueba": None,
                "ValorMedido": None,
                "Percentil": None,
                "ClasificacionFuncional": None,
                "Peso_kg": row.get("peso_kg"),
                "IMC": row.get("imc"),
                "GrasaCorporal_pct": row.get("grasa_corporal_pct"),
                "MasaMuscular_kg": row.get("masa_muscular_kg"),
                "AguaCorporal_pct": row.get("agua_corporal_pct"),
                "GrasaVisceral": row.get("grasa_visceral"),
                "MetabolismoBasal": row.get("metabolismo_basal"),
                "MusculoRelativo_pct": row.get("musculo_rel_pct"),
                "Clasif_IMC": row.get("clasif_imc"),
                "Clasif_Grasa": row.get("clasif_grasa"),
                "Clasif_Agua": row.get("clasif_agua"),
                "Clasif_Visceral": row.get("clasif_visceral"),
                "Clasif_Musculo": row.get("clasif_musculo"),
                "DiagnosticoCorporal": row.get("diagnostico_corporal"),
                "SugerenciaCorporal": row.get("sugerencia_corporal"),
                "MotivosCorporal": row.get("motivos_corporal"),
                "Observaciones": row.get("observaciones")
            })

    if df_eval is not None and not df_eval.empty:
        df_f = df_eval.copy()
        if "fecha" in df_f.columns:
            df_f["fecha"] = pd.to_datetime(df_f["fecha"], errors="coerce")

        for _, row in df_f.iterrows():
            filas.append({
                "PacienteID": paciente_id,
                "Paciente": nombre,
                "Sexo": sexo,
                "Talla_m": talla_m,
                "Fecha": row.get("fecha"),
                "TipoRegistro": "Evaluacion_Funcional",
                "Prueba": row.get("prueba"),
                "ValorMedido": row.get("valor_medido"),
                "Percentil": row.get("percentil"),
                "ClasificacionFuncional": row.get("clasificacion"),
                "Peso_kg": None,
                "IMC": None,
                "GrasaCorporal_pct": None,
                "MasaMuscular_kg": None,
                "AguaCorporal_pct": None,
                "GrasaVisceral": None,
                "MetabolismoBasal": None,
                "MusculoRelativo_pct": None,
                "Clasif_IMC": None,
                "Clasif_Grasa": None,
                "Clasif_Agua": None,
                "Clasif_Visceral": None,
                "Clasif_Musculo": None,
                "DiagnosticoCorporal": None,
                "SugerenciaCorporal": None,
                "MotivosCorporal": None,
                "Observaciones": None
            })

    df_final = pd.DataFrame(filas)

    if not df_final.empty:
        df_final["Fecha"] = pd.to_datetime(df_final["Fecha"], errors="coerce")
        df_final = df_final.sort_values(["Paciente", "Fecha", "TipoRegistro"], ascending=[True, True, True]).reset_index(drop=True)

    return df_final


# =========================================================
# EXPORTACIÓN EXCEL
# =========================================================
def _formatear_hoja_excel(ws):
    if ws.max_row == 0 or ws.max_column == 0:
        return

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    fill_header = PatternFill(fill_type="solid", fgColor="1F4E78")
    font_header = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            try:
                cell_value = "" if cell.value is None else str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except Exception:
                pass

        ancho = min(max(max_length + 2, 12), 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def generar_excel_paciente(ficha, df_peso, df_inbody, df_eval):
    output = BytesIO()

    ficha_df = pd.DataFrame([{
        "Paciente": ficha["nombre"],
        "PacienteID_Ficha": ficha["id"],
        "Sexo": ficha["sexo"],
        "Talla_m": ficha["talla_m"],
        "CantidadEvaluaciones": ficha["cantidad_evaluaciones"],
        "UltimaFechaEvaluacion": ficha["ultima_fecha"],
        "UltimaClasificacion": ficha["ultima_clasificacion"],
        "UltimaPrueba": ficha["ultima_prueba"],
        "FechaExportacion": datetime.now().strftime("%Y-%m-%d %H:%M")
    }])

    df_peso_export = agregar_identificacion_paciente(df_peso, ficha, "Peso_IMC")
    df_inbody_enriquecido = enriquecer_historial_corporal(
        df_inbody,
        str(ficha["sexo"]).strip().lower(),
        ficha["talla_m"]
    )
    df_inbody_export = agregar_identificacion_paciente(df_inbody_enriquecido, ficha, "Composicion_Corporal")
    df_eval_export = agregar_identificacion_paciente(df_eval, ficha, "Evaluaciones")

    df_analisis = generar_df_analisis_cientifico(
        ficha=ficha,
        df_peso=df_peso,
        df_inbody=df_inbody,
        df_eval=df_eval
    )

    if not df_analisis.empty:
        df_analisis = df_analisis.rename(columns={"PacienteID": "PacienteID_Ficha"})

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        preparar_df_exportacion(ficha_df).to_excel(writer, sheet_name="00_Ficha", index=False)
        preparar_df_exportacion(df_peso_export).to_excel(writer, sheet_name="01_Peso_IMC", index=False)
        preparar_df_exportacion(df_inbody_export).to_excel(writer, sheet_name="02_Composicion", index=False)
        preparar_df_exportacion(df_eval_export).to_excel(writer, sheet_name="03_Evaluaciones", index=False)
        preparar_df_exportacion(df_analisis).to_excel(writer, sheet_name="04_Analisis", index=False)

        workbook = writer.book
        for nombre_hoja in workbook.sheetnames:
            _formatear_hoja_excel(workbook[nombre_hoja])

    output.seek(0)
    return output


# =========================================================
# EXPORTACIÓN PDF
# =========================================================
def _texto_seguro(valor):
    if pd.isna(valor):
        return "-"
    valor = str(valor).strip()
    return valor if valor else "-"


def _df_para_pdf(df):
    if df is None or df.empty:
        return pd.DataFrame()

    out = df.copy()

    for col in out.columns:
        if "fecha" in col.lower() or "created" in col.lower():
            out[col] = pd.to_datetime(out[col], errors="coerce").dt.strftime("%Y-%m-%d")

    return out.fillna("-")


def _tabla_pdf_desde_df(df, columnas, titulos, anchos_cm, styles):
    if df is None or df.empty:
        return Paragraph("Sin datos.", styles["Normal"])

    df = df.copy()
    columnas_validas = [c for c in columnas if c in df.columns]
    titulos_validos = [titulos[i] for i, c in enumerate(columnas) if c in df.columns]
    anchos_validos = [anchos_cm[i] * cm for i, c in enumerate(columnas) if c in df.columns]

    data = [[Paragraph(f"<b>{t}</b>", styles["TablaHeader"]) for t in titulos_validos]]

    for _, row in df.iterrows():
        fila = []
        for c in columnas_validas:
            txt = _texto_seguro(row.get(c))
            fila.append(Paragraph(txt, styles["TablaBody"]))
        data.append(fila)

    tabla = Table(data, colWidths=anchos_validos, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7.5),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C9D2DF")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#F7FAFC")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return tabla


def generar_pdf_paciente(ficha, df_peso, df_inbody, df_eval):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TituloMain", parent=styles["Heading1"], fontSize=18, leading=22, spaceAfter=10))
    styles.add(ParagraphStyle(name="SubTitulo", parent=styles["Heading2"], fontSize=12, leading=15, spaceBefore=8, spaceAfter=6))
    styles.add(ParagraphStyle(name="TablaHeader", parent=styles["Normal"], fontSize=8, textColor=colors.white))
    styles.add(ParagraphStyle(name="TablaBody", parent=styles["Normal"], fontSize=7.5, leading=9))
    styles.add(ParagraphStyle(name="Caja", parent=styles["Normal"], fontSize=9, leading=12))

    story = []

    story.append(Paragraph("Reporte completo del paciente", styles["TituloMain"]))
    story.append(Paragraph(f"<b>Paciente:</b> {_texto_seguro(ficha['nombre'])}", styles["Caja"]))
    story.append(Paragraph(f"<b>Sexo:</b> {_texto_seguro(ficha['sexo'])}", styles["Caja"]))
    story.append(Paragraph(f"<b>Talla:</b> {_texto_seguro(ficha['talla_m'])} m", styles["Caja"]))
    story.append(Paragraph(f"<b>Fecha del reporte:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Caja"]))
    story.append(Spacer(1, 0.25 * cm))

    df_peso_pdf = _df_para_pdf(df_peso)
    df_inbody_pdf = _df_para_pdf(df_inbody)
    if not df_inbody_pdf.empty:
        df_inbody_pdf = enriquecer_historial_corporal(
            df_inbody_pdf,
            str(ficha["sexo"]).strip().lower(),
            ficha["talla_m"]
        )
        df_inbody_pdf = _df_para_pdf(df_inbody_pdf)

    df_eval_pdf = _df_para_pdf(df_eval)

    if df_inbody_pdf is not None and not df_inbody_pdf.empty:
        ultimo_corporal = df_inbody_pdf.copy()
        if "fecha" in ultimo_corporal.columns:
            ultimo_corporal["fecha_orden"] = pd.to_datetime(ultimo_corporal["fecha"], errors="coerce")
            ultimo_corporal = ultimo_corporal.sort_values("fecha_orden", ascending=False)
        ult = ultimo_corporal.iloc[0]

        story.append(Paragraph("Diagnóstico corporal actual", styles["SubTitulo"]))
        story.append(Paragraph(f"<b>Diagnóstico:</b> {_texto_seguro(ult.get('diagnostico_corporal'))}", styles["Caja"]))
        story.append(Paragraph(f"<b>Sugerencia:</b> {_texto_seguro(ult.get('sugerencia_corporal'))}", styles["Caja"]))
        story.append(Paragraph(f"<b>Motivos:</b> {_texto_seguro(ult.get('motivos_corporal'))}", styles["Caja"]))
        story.append(Spacer(1, 0.25 * cm))

    story.append(Paragraph("Historial de peso e IMC", styles["SubTitulo"]))
    story.append(_tabla_pdf_desde_df(
        df_peso_pdf,
        columnas=["fecha", "peso_kg", "imc"],
        titulos=["Fecha", "Peso (kg)", "IMC"],
        anchos_cm=[4, 4, 3],
        styles=styles
    ))
    story.append(Spacer(1, 0.35 * cm))

    story.append(Paragraph("Historial de composición corporal", styles["SubTitulo"]))
    story.append(_tabla_pdf_desde_df(
        df_inbody_pdf,
        columnas=[
            "fecha",
            "peso_kg",
            "imc",
            "grasa_corporal_pct",
            "masa_muscular_kg",
            "agua_corporal_pct",
            "grasa_visceral",
            "diagnostico_corporal"
        ],
        titulos=[
            "Fecha",
            "Peso",
            "IMC",
            "% Grasa",
            "Músculo kg",
            "% Agua",
            "Visceral",
            "Diagnóstico"
        ],
        anchos_cm=[2.8, 2.1, 1.7, 2.2, 2.4, 2.1, 1.8, 4.1],
        styles=styles
    ))
    story.append(Spacer(1, 0.35 * cm))

    story.append(Paragraph("Evaluaciones funcionales", styles["SubTitulo"]))
    story.append(_tabla_pdf_desde_df(
        df_eval_pdf,
        columnas=["fecha", "prueba", "valor_medido", "percentil", "clasificacion"],
        titulos=["Fecha", "Prueba", "Valor", "Percentil", "Clasificación"],
        anchos_cm=[2.8, 5.7, 2.2, 2.0, 3.5],
        styles=styles
    ))

    doc.build(story)
    buffer.seek(0)
    return buffer


# =========================================================
# UTILIDADES CLÍNICAS - FUNCIONALES
# =========================================================
def clasificar_percentil(percentil):
    if percentil is None:
        return "Sin clasificar"
    if percentil < 10:
        return "Muy bajo"
    if percentil < 25:
        return "Bajo"
    if percentil < 50:
        return "Ligeramente bajo"
    if percentil < 75:
        return "Normal"
    if percentil < 90:
        return "Bueno"
    return "Muy bueno"


def color_clasificacion(clasificacion):
    colores = {
        "Muy bajo": "#d32f2f",
        "Bajo": "#f57c00",
        "Ligeramente bajo": "#fbc02d",
        "Normal": "#388e3c",
        "Bueno": "#1976d2",
        "Muy bueno": "#00796b",
        "Sin clasificar": "#757575"
    }
    return colores.get(clasificacion, "#757575")


def rango_percentilar(percentil):
    if percentil is None:
        return "Sin rango"
    if percentil < 3:
        return "Menor a P3"
    if percentil < 10:
        return "Entre P3 y P10"
    if percentil < 25:
        return "Entre P10 y P25"
    if percentil < 50:
        return "Entre P25 y P50"
    if percentil < 75:
        return "Entre P50 y P75"
    if percentil < 90:
        return "Entre P75 y P90"
    if percentil < 97:
        return "Entre P90 y P97"
    return "Mayor a P97"


def interpretacion_clinica(clasificacion):
    textos = {
        "Muy bajo": "Resultado marcadamente por debajo del rango funcional esperado.",
        "Bajo": "Resultado por debajo del rango funcional esperado.",
        "Ligeramente bajo": "Resultado levemente inferior al rango esperado.",
        "Normal": "Resultado dentro del rango funcional esperado.",
        "Bueno": "Resultado superior al promedio esperado.",
        "Muy bueno": "Resultado claramente superior al rango esperado."
    }
    return textos.get(clasificacion, "Sin interpretación disponible.")


def interpolar_percentil(valor_medido, tabla_percentiles):
    puntos = sorted(tabla_percentiles.items(), key=lambda x: x[1])

    if valor_medido < puntos[0][1]:
        return float(puntos[0][0])

    if valor_medido > puntos[-1][1]:
        return float(puntos[-1][0])

    for i in range(len(puntos) - 1):
        p1, v1 = puntos[i]
        p2, v2 = puntos[i + 1]

        if v1 <= valor_medido <= v2:
            if v2 == v1:
                return float(p1)
            return float(p1 + (valor_medido - v1) * (p2 - p1) / (v2 - v1))

    return None


def grupo_edad_prension(edad):
    edad = int(edad)
    if edad >= 100:
        return "+100"
    inicio = max(20, min(95, 20 + ((edad - 20) // 5) * 5))
    fin = inicio + 4
    return f"{inicio}-{fin}"


def grupo_edad_silla(edad):
    edad = int(edad)
    if edad >= 85:
        return "+84"
    if 65 <= edad <= 69:
        return "65-69"
    if 70 <= edad <= 74:
        return "70-74"
    if 75 <= edad <= 79:
        return "75-79"
    if 80 <= edad <= 84:
        return "80-84"
    return None


def calcular_resultado(prueba, sexo, edad, altura, valor_medido):
    sexo = str(sexo).strip()
    edad = int(edad)
    valor_medido = float(valor_medido)

    if prueba == "Caminata 6 minutos":
        altura = int(altura)
        altura_ref = min(TABLA_CAMINATA_6M.keys(), key=lambda x: abs(x - altura))
        edad_ref = min(TABLA_CAMINATA_6M[altura_ref].keys(), key=lambda x: abs(x - edad))

        percentiles = TABLA_CAMINATA_6M[altura_ref][edad_ref]
        percentil_estimado = interpolar_percentil(valor_medido, percentiles)

        if percentil_estimado is None:
            return None, "Sin clasificar", "-", "-", "-"

        percentil_estimado = round(percentil_estimado, 1)
        clasificacion = clasificar_percentil(percentil_estimado)
        referencia_p50 = percentiles[50]

        return (
            percentil_estimado,
            clasificacion,
            f"{referencia_p50} m",
            f"Altura ref.: {altura_ref} cm",
            f"Edad ref.: {edad_ref} años"
        )

    if prueba == "Prensión manual":
        grupo = grupo_edad_prension(edad)
        percentiles = TABLA_PRENSION[sexo][grupo]
        percentil_estimado = interpolar_percentil(valor_medido, percentiles)

        if percentil_estimado is None:
            return None, "Sin clasificar", "-", "-", "-"

        percentil_estimado = round(percentil_estimado, 1)
        clasificacion = clasificar_percentil(percentil_estimado)
        referencia_p50 = percentiles[50]

        return (
            percentil_estimado,
            clasificacion,
            f"{referencia_p50} kg",
            "-",
            f"Grupo etario: {grupo}"
        )

    if prueba == "Levantarse de la silla":
        grupo = grupo_edad_silla(edad)

        if grupo is None:
            return None, "Sin clasificar", "-", "-", "Edad fuera del rango de la tabla (65+ años)"

        percentiles = TABLA_SILLA[sexo][grupo]
        percentil_estimado = interpolar_percentil(valor_medido, percentiles)

        if percentil_estimado is None:
            return None, "Sin clasificar", "-", "-", "-"

        percentil_estimado = round(percentil_estimado, 1)
        clasificacion = clasificar_percentil(percentil_estimado)
        referencia_p50 = percentiles[50]

        return (
            percentil_estimado,
            clasificacion,
            f"{referencia_p50} rep",
            "-",
            f"Grupo etario: {grupo}"
        )

    return None, "Sin clasificar", "-", "-", "-"


# =========================================================
# UI
# =========================================================
st.title("Calculadora de Condición Física")

with st.expander("➕ Nuevo paciente"):
    nuevo_nombre = st.text_input("Nombre del nuevo paciente", key="nuevo_nombre_alta")
    nuevo_sexo = st.selectbox("Sexo del nuevo paciente", ["hombre", "mujer"], key="nuevo_sexo_alta")
    nueva_talla = st.number_input(
        "Talla (m)",
        min_value=0.50,
        max_value=2.50,
        value=1.70,
        step=0.01,
        format="%.2f",
        key="nueva_talla_alta"
    )

    if st.button("Guardar paciente", key="btn_guardar_paciente"):
        if not nuevo_nombre.strip():
            st.warning("Ingresá el nombre del paciente.")
        else:
            try:
                guardar_paciente(nuevo_nombre, nuevo_sexo, nueva_talla)
                st.success("Paciente agregado correctamente.")
                st.rerun()
            except Exception as e:
                st.error(f"Error al guardar paciente: {e}")

pacientes = obtener_pacientes()
opciones_pacientes = [p["nombre"] for p in pacientes]

if not opciones_pacientes:
    st.warning("No hay pacientes cargados. Agregá uno desde '➕ Nuevo paciente'.")
    st.stop()

# =========================================================
# ENCABEZADO
# =========================================================
top1, top2, top3 = st.columns([2, 1, 1])

with top1:
    paciente_nombre = st.selectbox(
        "Seleccionar paciente",
        opciones_pacientes,
        key="selector_paciente"
    )

paciente_actual = next((p for p in pacientes if p["nombre"] == paciente_nombre), None)
paciente_id = paciente_actual["id"] if paciente_actual else None

# =========================================================
# FICHA + DATAFRAMES BASE
# =========================================================
ficha = obtener_ficha_paciente(paciente_nombre)
df_peso_export = obtener_historial_peso(paciente_id) if paciente_id is not None else pd.DataFrame()
df_inbody_export = obtener_historial_inbody(paciente_id) if paciente_id is not None else pd.DataFrame()
df_eval_export = obtener_historial_paciente(paciente_nombre) if paciente_nombre else pd.DataFrame()

excel_buffer = generar_excel_paciente(ficha, df_peso_export, df_inbody_export, df_eval_export)
pdf_buffer = generar_pdf_paciente(ficha, df_peso_export, df_inbody_export, df_eval_export)

with top2:
    st.download_button(
        label="Descargar Excel",
        data=excel_buffer.getvalue(),
        file_name=f"{paciente_nombre.replace(' ', '_')}_reporte.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"descargar_excel_{paciente_id}"
    )

with top3:
    st.download_button(
        label="Descargar PDF",
        data=pdf_buffer.getvalue(),
        file_name=f"{paciente_nombre.replace(' ', '_')}_reporte.pdf",
        mime="application/pdf",
        key=f"descargar_pdf_{paciente_id}"
    )

st.divider()

# =========================================================
# TARJETAS RESUMEN
# =========================================================
k1, k2, k3, k4 = st.columns(4)

with k1:
    with st.container(border=True):
        st.markdown("#### Paciente")
        st.write(f"**Nombre:** {ficha['nombre']}")
        st.write(f"**Sexo:** {str(ficha['sexo']).capitalize() if ficha['sexo'] != '-' else '-'}")
        if ficha["talla_m"] is not None:
            st.write(f"**Talla:** {float(ficha['talla_m']):.2f} m")
        else:
            st.write("**Talla:** -")

with k2:
    with st.container(border=True):
        st.markdown("#### Peso / IMC")
        if not df_peso_export.empty:
            df_peso_tmp = df_peso_export.copy()
            df_peso_tmp["fecha"] = pd.to_datetime(df_peso_tmp["fecha"], errors="coerce")
            df_peso_tmp = df_peso_tmp.dropna(subset=["fecha"]).sort_values("fecha", ascending=False)
            ultimo_peso = df_peso_tmp.iloc[0]
            st.write(f"**Peso:** {float(ultimo_peso['peso_kg']):.1f} kg")
            st.write(f"**IMC:** {float(ultimo_peso['imc']):.2f}")
            st.write(f"**Fecha:** {ultimo_peso['fecha'].strftime('%d-%m-%Y')}")
        else:
            st.write("Sin registros")

with k3:
    with st.container(border=True):
        st.markdown("#### Estado corporal")
        if not df_inbody_export.empty and ficha["talla_m"] is not None:
            df_corporal_tmp = enriquecer_historial_corporal(
                df_inbody_export,
                str(ficha["sexo"]).strip().lower(),
                ficha["talla_m"]
            )
            df_corporal_tmp["fecha"] = pd.to_datetime(df_corporal_tmp["fecha"], errors="coerce")
            df_corporal_tmp = df_corporal_tmp.dropna(subset=["fecha"]).sort_values("fecha", ascending=False)
            ultimo_corporal = df_corporal_tmp.iloc[0]

            st.write(f"**Estado:** {ultimo_corporal.get('diagnostico_corporal', '-')}")
            st.write(f"**Grasa:** {ultimo_corporal.get('clasif_grasa', '-')}")
            st.write(f"**Músculo:** {ultimo_corporal.get('clasif_musculo', '-')}")
        else:
            st.write("Sin registros")

with k4:
    with st.container(border=True):
        st.markdown("#### Última evaluación")
        st.write(f"**Fecha:** {ficha['ultima_fecha']}")
        st.write(f"**Prueba:** {ficha['ultima_prueba']}")
        st.write(f"**Clasificación:** {ficha['ultima_clasificacion']}")

st.divider()

# =========================================================
# ZONA PRINCIPAL DE TRABAJO
# =========================================================
left, right = st.columns([1, 1])

with left:
    st.markdown("## Peso e IMC")

    with st.container(border=True):
        if ficha["talla_m"] is None or float(ficha["talla_m"]) <= 0:
            st.warning("Este paciente no tiene una talla válida cargada en la tabla pacientes.")
        else:
            col_p1, col_p2, col_p3 = st.columns(3)

            with col_p1:
                fecha_peso = st.date_input(
                    "Fecha de peso",
                    value=date.today(),
                    key=f"fecha_peso_{paciente_id}"
                )

            with col_p2:
                peso_kg = st.number_input(
                    "Peso (kg)",
                    min_value=0.0,
                    max_value=300.0,
                    value=70.0,
                    step=0.1,
                    format="%.1f",
                    key=f"peso_kg_{paciente_id}"
                )

            with col_p3:
                imc_calculado = round(float(peso_kg) / (float(ficha["talla_m"]) ** 2), 2)
                clasificacion_imc, color_imc = clasificar_imc(imc_calculado)
                st.markdown(f"**IMC:** {imc_calculado:.2f}")
                st.markdown(f"**Clasificación:** {color_imc} {clasificacion_imc}")

            if st.button("Guardar peso", key=f"btn_guardar_peso_{paciente_id}"):
                try:
                    guardar_peso(
                        paciente_id=paciente_id,
                        fecha_medicion=fecha_peso,
                        peso_kg=peso_kg,
                        talla_m=ficha["talla_m"]
                    )
                    st.success("Peso guardado correctamente.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al guardar peso: {e}")

    st.markdown("## Composición corporal")

    with st.container(border=True):
        if ficha["talla_m"] is None or float(ficha["talla_m"]) <= 0:
            st.warning("Para cargar composición corporal primero hay que tener una talla válida en el paciente.")
        else:
            sexo_corporal = str(ficha["sexo"]).strip().lower()

            col_c1, col_c2 = st.columns(2)

            with col_c1:
                fecha_inbody = st.date_input("Fecha estudio", value=date.today(), key=f"inbody_fecha_{paciente_id}")
                peso_inbody = st.number_input("Peso (kg)", min_value=0.0, max_value=300.0, step=0.1, key=f"inbody_peso_{paciente_id}")
                imc_inbody_calc = round(float(peso_inbody) / (float(ficha["talla_m"]) ** 2), 2) if peso_inbody and ficha["talla_m"] else 0.0
                st.markdown(f"**IMC calculado:** {imc_inbody_calc:.2f}")
                grasa_pct = st.number_input("% grasa corporal", min_value=0.0, max_value=80.0, step=0.1, key=f"inbody_grasa_{paciente_id}")

            with col_c2:
                masa_muscular = st.number_input("Masa muscular (kg)", min_value=0.0, max_value=100.0, step=0.1, key=f"inbody_musculo_{paciente_id}")
                agua_pct = st.number_input("% agua corporal", min_value=0.0, max_value=100.0, step=0.1, key=f"inbody_agua_{paciente_id}")
                grasa_visceral = st.number_input("Grasa visceral", min_value=0.0, max_value=30.0, step=0.1, key=f"inbody_visceral_{paciente_id}")
                metabolismo = st.number_input("Metabolismo basal", min_value=0.0, max_value=4000.0, step=10.0, key=f"inbody_metabolismo_{paciente_id}")

            observaciones_inbody = st.text_area("Observaciones", key=f"inbody_obs_{paciente_id}")

            resultado_corporal = evaluar_perfil_morfofuncional(
                sexo=sexo_corporal,
                peso_kg=peso_inbody,
                talla_m=ficha["talla_m"],
                grasa_pct=grasa_pct,
                masa_muscular_kg=masa_muscular,
                agua_pct=agua_pct,
                grasa_visceral=grasa_visceral
            )

            bg_estado, fg_estado = color_estado_corporal(resultado_corporal["estado"])

            st.markdown(
                f"""
                <div class="result-card" style="background-color:{bg_estado}; color:{fg_estado};">
                    Diagnóstico corporal: {resultado_corporal["estado"]}
                </div>
                """,
                unsafe_allow_html=True
            )

            col_r1, col_r2, col_r3 = st.columns(3)
            with col_r1:
                st.write(f"**IMC:** {resultado_corporal['imc'] if resultado_corporal['imc'] is not None else '-'}")
                st.write(f"**Clasificación IMC:** {resultado_corporal['clasif_imc']}")
                st.write(f"**% grasa:** {resultado_corporal['clasif_grasa']}")

            with col_r2:
                st.write(f"**% agua corporal:** {resultado_corporal['clasif_agua']}")
                st.write(f"**Grasa visceral:** {resultado_corporal['clasif_visceral']}")
                st.write(f"**Músculo relativo %:** {resultado_corporal['musculo_rel_pct'] if resultado_corporal['musculo_rel_pct'] is not None else '-'}")

            with col_r3:
                st.write(f"**Clasificación muscular:** {resultado_corporal['clasif_musculo']}")
                st.write(f"**Metabolismo basal:** {metabolismo if metabolismo is not None else '-'}")
                st.write(f"**Sexo de referencia:** {sexo_corporal.capitalize() if sexo_corporal else '-'}")

            motivos_texto = resultado_corporal["motivos"] if resultado_corporal["motivos"] else ["Sin hallazgos relevantes"]
            motivos_html = "".join([f"<li>{m}</li>" for m in motivos_texto])

            st.markdown(
                f"""
                <div class="motivo-box">
                    <b>Motivos:</b>
                    <ul style="margin-top:8px; margin-bottom:0;">
                        {motivos_html}
                    </ul>
                </div>
                """,
                unsafe_allow_html=True
            )

            st.markdown(
                f"""
                <div class="reco-box">
                    <b>Sugerencia:</b><br>
                    {resultado_corporal["recomendacion"]}
                </div>
                """,
                unsafe_allow_html=True
            )

            if st.button("Guardar composición corporal", key=f"guardar_inbody_{paciente_id}"):
                try:
                    guardar_inbody(
                        paciente_id=paciente_id,
                        fecha_estudio=fecha_inbody,
                        peso_kg=peso_inbody,
                        talla_m=ficha["talla_m"],
                        grasa_corporal_pct=grasa_pct,
                        masa_muscular_kg=masa_muscular,
                        agua_corporal_pct=agua_pct,
                        grasa_visceral=grasa_visceral,
                        metabolismo_basal=metabolismo,
                        observaciones=observaciones_inbody
                    )
                    st.success("Composición corporal guardada correctamente")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al guardar composición corporal: {e}")

with right:
    st.markdown("## Evaluación funcional")

    with st.container(border=True):
        paciente_sexo_guardado = next((p["sexo"] for p in pacientes if p["nombre"] == paciente_nombre), None)

        prueba = st.selectbox(
            "Seleccionar prueba",
            ["Caminata 6 minutos", "Prensión manual", "Levantarse de la silla"],
            key="selector_prueba"
        )

        sexo_default = "Hombre"
        if paciente_sexo_guardado == "mujer":
            sexo_default = "Mujer"
        elif paciente_sexo_guardado == "hombre":
            sexo_default = "Hombre"

        sexo = st.selectbox(
            "Sexo",
            ["Hombre", "Mujer"],
            index=0 if sexo_default == "Hombre" else 1,
            key="selector_sexo"
        )

        altura = None
        valor_medido = None

        if prueba == "Caminata 6 minutos":
            edad = st.selectbox("Edad", list(range(40, 81)), index=20, key="edad_caminata")
            altura = st.selectbox("Altura (cm)", [150, 160, 170, 180, 190], index=2, key="altura_caminata")
            valor_medido = st.number_input(
                "Distancia caminada (metros)",
                min_value=0.0,
                max_value=2000.0,
                value=600.0,
                step=1.0,
                format="%.2f",
                key="valor_caminata"
            )

        elif prueba == "Prensión manual":
            edad = st.selectbox("Edad", list(range(20, 101)), index=45, key="edad_prension")
            valor_medido = st.number_input(
                "Fuerza de prensión (kg)",
                min_value=0.0,
                max_value=100.0,
                value=25.0,
                step=0.1,
                format="%.1f",
                key="valor_prension"
            )

        elif prueba == "Levantarse de la silla":
            edad = st.selectbox("Edad", list(range(65, 101)), index=10, key="edad_silla")
            valor_medido = st.number_input(
                "Cantidad de repeticiones",
                min_value=0.0,
                max_value=60.0,
                value=12.0,
                step=1.0,
                format="%.0f",
                key="valor_silla"
            )

        percentil, clasificacion, referencia_p50, referencia_altura, referencia_edad = calcular_resultado(
            prueba=prueba,
            sexo=sexo,
            edad=edad,
            altura=altura,
            valor_medido=valor_medido
        )

        color = color_clasificacion(clasificacion)

        st.markdown(
            f"""
            <div style="
                background-color:{color};
                color:white;
                padding:10px 12px;
                border-radius:8px;
                text-align:center;
                font-size:18px;
                font-weight:600;
                margin-top:18px;
                margin-bottom:10px;
            ">
                {clasificacion}
            </div>
            """,
            unsafe_allow_html=True
        )

        st.markdown(
            f"""
            <div style="
                background-color:#dff0e6;
                color:#1b5e20;
                padding:8px 12px;
                border-radius:8px;
                font-size:15px;
                margin-bottom:14px;
            ">
                Percentil estimado: <b>P{percentil if percentil is not None else "-"}</b>
            </div>
            """,
            unsafe_allow_html=True
        )

        st.write(f"**Rango percentilar:** {rango_percentilar(percentil)}")
        st.write(f"**Referencia P50:** {referencia_p50}")

        if referencia_altura != "-":
            st.write(f"**Referencia de altura:** {referencia_altura}")

        if referencia_edad != "-":
            st.write(f"**Referencia etaria:** {referencia_edad}")

        st.write(f"**Interpretación clínica:** {interpretacion_clinica(clasificacion)}")

        if st.button("Guardar evaluación", key="btn_guardar_evaluacion"):
            if not paciente_nombre:
                st.warning("Seleccioná un paciente antes de guardar.")
            elif percentil is None:
                st.warning("No se pudo calcular el percentil.")
            else:
                try:
                    guardar_evaluacion(
                        paciente_nombre=paciente_nombre,
                        sexo=sexo,
                        edad=edad,
                        prueba=prueba,
                        valor_medido=valor_medido,
                        percentil=percentil,
                        clasificacion=clasificacion
                    )
                    st.success("Evaluación guardada correctamente.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al guardar: {e}")

st.divider()

# =========================================================
# HISTORIALES
# =========================================================
h1, h2 = st.columns([1, 1])

with h1:
    st.markdown("## Historial corporal")

    df_inbody = obtener_historial_inbody(paciente_id)
    if not df_inbody.empty:
        df_inbody = enriquecer_historial_corporal(
            df_inbody,
            str(ficha["sexo"]).strip().lower(),
            ficha["talla_m"]
        )
        df_inbody["fecha"] = pd.to_datetime(df_inbody["fecha"], errors="coerce")
        df_inbody = df_inbody.dropna(subset=["fecha"]).sort_values("fecha", ascending=False)

        df_inbody_mostrar = agregar_identificacion_paciente(df_inbody, ficha, "Composicion_Corporal")
        df_inbody_mostrar["fecha"] = pd.to_datetime(df_inbody_mostrar["fecha"], errors="coerce").dt.strftime("%Y-%m-%d")

        columnas_inbody = [
            "Paciente",
            "PacienteID_Ficha",
            "Sexo",
            "Talla_m",
            "Origen",
            "fecha",
            "peso_kg",
            "imc",
            "grasa_corporal_pct",
            "masa_muscular_kg",
            "agua_corporal_pct",
            "grasa_visceral",
            "metabolismo_basal",
            "diagnostico_corporal",
            "sugerencia_corporal",
            "motivos_corporal",
            "observaciones"
        ]
        columnas_inbody = [c for c in columnas_inbody if c in df_inbody_mostrar.columns]

        st.dataframe(
            df_inbody_mostrar[columnas_inbody],
            use_container_width=True,
            hide_index=True
        )
    else:
        st.info("Sin historial corporal.")

with h2:
    st.markdown("## Historial funcional")

    df_historial = obtener_historial_paciente(paciente_nombre)

    if not df_historial.empty:
        prueba_filtro = st.selectbox(
            "Filtrar historial por prueba",
            options=["Todas", "Caminata 6 minutos", "Prensión manual", "Levantarse de la silla"],
            index=0,
            key="filtro_historial_prueba"
        )

        if prueba_filtro == "Todas":
            df_historial_filtrado = df_historial.copy()
        else:
            df_historial_filtrado = df_historial[
                df_historial["prueba"].astype(str).str.strip() == prueba_filtro
            ].copy()

        columnas_mostrar = ["id", "fecha", "prueba", "valor_medido", "percentil", "clasificacion"]
        columnas_existentes = [c for c in columnas_mostrar if c in df_historial_filtrado.columns]
        df_historial_mostrar = df_historial_filtrado[columnas_existentes].copy()

        if "fecha" in df_historial_mostrar.columns:
            df_historial_mostrar["fecha"] = pd.to_datetime(
                df_historial_mostrar["fecha"],
                errors="coerce"
            ).dt.strftime("%Y-%m-%d")

        df_historial_mostrar = df_historial_mostrar.sort_values(by="fecha", ascending=False)

        st.markdown("**Fecha | Prueba | Valor | Percentil | Clasificación | Eliminar**")

        for _, row in df_historial_mostrar.iterrows():
            c1, c2, c3, c4, c5, c6 = st.columns([1, 2, 1, 1, 1, 0.5])

            c1.write(row.get("fecha", ""))
            c2.write(row.get("prueba", ""))
            c3.write(row.get("valor_medido", ""))
            c4.write(row.get("percentil", ""))
            c5.write(row.get("clasificacion", ""))

            if c6.button("🗑", key=f"del_{row['id']}"):
                try:
                    eliminar_evaluacion(row["id"])
                    st.success("Evaluación eliminada.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al eliminar: {e}")
    else:
        st.info("Sin historial funcional.")

st.divider()

# =========================================================
# GRÁFICOS
# =========================================================
st.markdown("## Gráficos")

g1, g2 = st.columns([1, 1])

with g1:
    st.markdown("### Evolución de peso e IMC")

    if paciente_id is not None:
        df_peso = obtener_historial_peso(paciente_id)

        if not df_peso.empty:
            df_peso["fecha"] = pd.to_datetime(df_peso["fecha"], errors="coerce")
            df_peso["peso_kg"] = pd.to_numeric(df_peso["peso_kg"], errors="coerce")
            df_peso["imc"] = pd.to_numeric(df_peso["imc"], errors="coerce")
            df_peso = df_peso.dropna(subset=["fecha", "peso_kg", "imc"]).sort_values("fecha", ascending=True)

            grafico_doble = alt.Chart(df_peso).transform_fold(
                ["peso_kg", "imc"],
                as_=["variable", "valor"]
            ).mark_line(point=True).encode(
                x=alt.X("yearmonthdate(fecha):T", title="Fecha"),
                y=alt.Y("valor:Q", title="Valor"),
                color=alt.Color("variable:N", title="Serie"),
                tooltip=[
                    alt.Tooltip("yearmonthdate(fecha):T", title="Fecha"),
                    alt.Tooltip("variable:N", title="Serie"),
                    alt.Tooltip("valor:Q", title="Valor", format=".2f")
                ]
            ).properties(height=320)

            st.altair_chart(grafico_doble, use_container_width=True)
        else:
            st.info("Sin datos de peso / IMC.")

with g2:
    st.markdown("### Evolución del percentil funcional")

    if paciente_nombre:
        df_historial = obtener_historial_paciente(paciente_nombre)

        if not df_historial.empty and {"fecha", "percentil", "prueba"}.issubset(df_historial.columns):
            df_graf_base = df_historial.copy()
            df_graf_base["fecha"] = pd.to_datetime(df_graf_base["fecha"], errors="coerce").dt.date
            df_graf_base["percentil"] = pd.to_numeric(df_graf_base["percentil"], errors="coerce")
            df_graf_base["prueba"] = df_graf_base["prueba"].astype(str).str.strip()
            df_graf_base = df_graf_base.dropna(subset=["fecha", "percentil", "prueba"])

            prueba_grafico = st.selectbox(
                "Prueba para gráfico",
                options=["Caminata 6 minutos", "Prensión manual", "Levantarse de la silla"],
                key="selector_grafico_prueba"
            )

            df_prueba = df_graf_base[df_graf_base["prueba"] == prueba_grafico].copy()

            if not df_prueba.empty:
                df_prueba = (
                    df_prueba.groupby("fecha", as_index=False)["percentil"]
                    .mean()
                    .sort_values("fecha")
                )

                df_prueba["Etiqueta"] = df_prueba["percentil"].apply(lambda x: f"P{round(x, 1)}")

                linea = alt.Chart(df_prueba).mark_line(point=False).encode(
                    x=alt.X(
                        "yearmonthdate(fecha):T",
                        title="Fecha",
                        axis=alt.Axis(format="%d-%m-%Y")
                    ),
                    y=alt.Y("percentil:Q", title="Percentil")
                )

                puntos = alt.Chart(df_prueba).mark_circle(size=90).encode(
                    x=alt.X("yearmonthdate(fecha):T"),
                    y=alt.Y("percentil:Q")
                )

                etiquetas = alt.Chart(df_prueba).mark_text(
                    dy=-12,
                    fontSize=12
                ).encode(
                    x=alt.X("yearmonthdate(fecha):T"),
                    y=alt.Y("percentil:Q"),
                    text="Etiqueta:N"
                )

                grafico = (linea + puntos + etiquetas).properties(height=320)
                st.altair_chart(grafico, use_container_width=True)

                ultimo = df_prueba["percentil"].iloc[-1]
                if len(df_prueba) >= 2:
                    anterior = df_prueba["percentil"].iloc[-2]
                    diferencia = round(ultimo - anterior, 1)

                    if diferencia > 0:
                        st.success(f"↑ Mejora de {diferencia} percentiles desde la evaluación anterior")
                    elif diferencia < 0:
                        st.warning(f"↓ Disminución de {abs(diferencia)} percentiles desde la evaluación anterior")
                    else:
                        st.info("Sin cambios respecto a la evaluación anterior")
            else:
                st.info("Sin datos funcionales para esa prueba.")
        else:
            st.info("Sin historial funcional para graficar.")
```

Reemplazá todo tu `app.py` por esto y probalo.
Si aparece un error, mandame **solo la captura del error** y te lo corrijo sobre eso.
import streamlit as st
import pandas as pd
import altair as alt
from datetime import datetime, date
from supabase import create_client, Client
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from io import BytesIO

# =========================================================
# CONFIG
# =========================================================
st.set_page_config(
    page_title="Calculadora de Condición Física",
    page_icon="💪",
    layout="wide"
)

st.markdown("""
<style>
[data-testid="stAppViewContainer"] {
    background-color: #f7f8fa;
}

[data-testid="stHeader"] {
    background: rgba(0, 0, 0, 0);
}

[data-testid="stToolbar"] {
    right: 1rem;
}

.block-container {
    padding-top: 1.5rem;
    padding-bottom: 2rem;
    max-width: 1400px;
}

.result-card {
    padding: 14px 16px;
    border-radius: 10px;
    font-size: 18px;
    font-weight: 700;
    margin-top: 8px;
    margin-bottom: 10px;
}

.motivo-box {
    background-color: #ffffff;
    border: 1px solid #e6e9ef;
    border-radius: 10px;
    padding: 12px 14px;
    margin-bottom: 12px;
}

.reco-box {
    background-color: #eef7ef;
    border: 1px solid #d4ead7;
    border-radius: 10px;
    padding: 12px 14px;
    margin-bottom: 14px;
}
</style>
""", unsafe_allow_html=True)

# =========================================================
# SUPABASE
# =========================================================
SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# =========================================================
# TABLAS NORMATIVAS
# =========================================================
TABLA_CAMINATA_6M = {
    150: {
        40: {2.5: 436, 10: 470, 25: 511, 50: 555, 75: 592, 90: 631, 97.5: 679},
        50: {2.5: 434, 10: 468, 25: 509, 50: 553, 75: 590, 90: 629, 97.5: 677},
        60: {2.5: 414, 10: 448, 25: 489, 50: 533, 75: 570, 90: 609, 97.5: 656},
        70: {2.5: 364, 10: 397, 25: 439, 50: 483, 75: 520, 90: 558, 97.5: 606},
        80: {2.5: 313, 10: 347, 25: 388, 50: 432, 75: 469, 90: 508, 97.5: 556},
    },
    160: {
        40: {2.5: 455, 10: 489, 25: 530, 50: 574, 75: 611, 90: 650, 97.5: 697},
        50: {2.5: 453, 10: 487, 25: 528, 50: 572, 75: 609, 90: 648, 97.5: 695},
        60: {2.5: 433, 10: 466, 25: 508, 50: 552, 75: 588, 90: 627, 97.5: 675},
        70: {2.5: 382, 10: 416, 25: 457, 50: 501, 75: 538, 90: 577, 97.5: 625},
        80: {2.5: 332, 10: 366, 25: 407, 50: 451, 75: 488, 90: 526, 97.5: 574},
    },
    170: {
        40: {2.5: 474, 10: 507, 25: 549, 50: 593, 75: 629, 90: 668, 97.5: 716},
        50: {2.5: 472, 10: 505, 25: 546, 50: 590, 75: 627, 90: 666, 97.5: 714},
        60: {2.5: 451, 10: 485, 25: 526, 50: 570, 75: 607, 90: 646, 97.5: 694},
        70: {2.5: 401, 10: 435, 25: 476, 50: 520, 75: 557, 90: 595, 97.5: 643},
        80: {2.5: 351, 10: 384, 25: 425, 50: 469, 75: 506, 90: 545, 97.5: 593},
    },
    180: {
        40: {2.5: 492, 10: 526, 25: 567, 50: 611, 75: 648, 90: 687, 97.5: 735},
        50: {2.5: 490, 10: 524, 25: 565, 50: 609, 75: 646, 90: 685, 97.5: 733},
        60: {2.5: 470, 10: 503, 25: 545, 50: 589, 75: 626, 90: 664, 97.5: 712},
        70: {2.5: 419, 10: 453, 25: 494, 50: 538, 75: 575, 90: 614, 97.5: 662},
        80: {2.5: 369, 10: 403, 25: 444, 50: 488, 75: 525, 90: 564, 97.5: 611},
    },
    190: {
        40: {2.5: 511, 10: 544, 25: 586, 50: 630, 75: 667, 90: 705, 97.5: 753},
        50: {2.5: 509, 10: 542, 25: 584, 50: 628, 75: 665, 90: 703, 97.5: 751},
        60: {2.5: 488, 10: 522, 25: 563, 50: 607, 75: 644, 90: 683, 97.5: 731},
        70: {2.5: 438, 10: 472, 25: 513, 50: 557, 75: 594, 90: 633, 97.5: 680},
        80: {2.5: 388, 10: 421, 25: 463, 50: 507, 75: 544, 90: 582, 97.5: 630},
    }
}

TABLA_PRENSION = {
    "Hombre": {
        "20-24": {5: 33.9, 10: 36.8, 20: 40.5, 30: 43.2, 40: 45.7, 50: 48.0, 60: 50.4, 70: 52.9, 80: 56.0, 90: 60.1, 95: 63.6},
        "25-29": {5: 35.5, 10: 38.5, 20: 42.1, 30: 44.8, 40: 47.1, 50: 49.3, 60: 51.5, 70: 53.9, 80: 56.7, 90: 60.7, 95: 64.0},
        "30-34": {5: 35.0, 10: 38.3, 20: 42.2, 30: 45.0, 40: 47.4, 50: 49.7, 60: 52.0, 70: 54.4, 80: 57.4, 90: 61.5, 95: 64.9},
        "35-39": {5: 33.8, 10: 37.3, 20: 41.5, 30: 44.5, 40: 47.1, 50: 49.5, 60: 51.9, 70: 54.4, 80: 57.5, 90: 61.8, 95: 65.3},
        "40-44": {5: 32.3, 10: 36.0, 20: 40.4, 30: 43.6, 40: 46.3, 50: 48.8, 60: 51.2, 70: 53.9, 80: 57.1, 90: 61.5, 95: 65.1},
        "45-49": {5: 30.6, 10: 34.4, 20: 39.0, 30: 42.3, 40: 45.1, 50: 47.6, 60: 50.2, 70: 52.9, 80: 56.2, 90: 60.7, 95: 64.4},
        "50-54": {5: 28.9, 10: 32.8, 20: 37.4, 30: 40.7, 40: 43.5, 50: 46.2, 60: 48.8, 70: 51.6, 80: 54.8, 90: 59.4, 95: 63.1},
        "55-59": {5: 27.2, 10: 31.0, 20: 35.6, 30: 38.9, 40: 41.7, 50: 44.4, 60: 47.0, 70: 49.8, 80: 53.1, 90: 57.7, 95: 61.4},
        "60-64": {5: 25.5, 10: 29.1, 20: 33.6, 30: 36.9, 40: 39.7, 50: 42.4, 60: 45.0, 70: 47.8, 80: 51.1, 90: 55.6, 95: 59.3},
        "65-69": {5: 23.7, 10: 27.2, 20: 31.5, 30: 34.7, 40: 37.5, 50: 40.1, 60: 42.8, 70: 45.6, 80: 48.8, 90: 53.2, 95: 56.8},
        "70-74": {5: 21.9, 10: 25.2, 20: 29.3, 30: 32.4, 40: 35.1, 50: 37.7, 60: 40.3, 70: 43.1, 80: 46.3, 90: 50.6, 95: 54.1},
        "75-79": {5: 20.0, 10: 23.1, 20: 27.0, 30: 29.9, 40: 32.5, 50: 35.1, 60: 37.6, 70: 40.3, 80: 43.5, 90: 47.7, 95: 51.1},
        "80-84": {5: 18.0, 10: 20.8, 20: 24.5, 30: 27.3, 40: 29.8, 50: 32.3, 60: 34.8, 70: 37.5, 80: 40.5, 90: 44.7, 95: 48.0},
        "85-89": {5: 15.9, 10: 18.5, 20: 21.9, 30: 24.6, 40: 27.0, 50: 29.4, 60: 31.8, 70: 34.4, 80: 37.4, 90: 41.5, 95: 44.6},
        "90-94": {5: 13.7, 10: 16.1, 20: 19.2, 30: 21.7, 40: 24.0, 50: 26.3, 60: 28.7, 70: 31.2, 80: 34.2, 90: 38.1, 95: 41.2},
        "95-99": {5: 11.3, 10: 13.5, 20: 16.4, 30: 18.8, 40: 20.9, 50: 23.1, 60: 25.4, 70: 27.9, 80: 30.8, 90: 34.6, 95: 37.5},
        "+100": {5: 8.8, 10: 10.8, 20: 13.5, 30: 15.7, 40: 17.8, 50: 19.8, 60: 22.0, 70: 24.5, 80: 27.2, 90: 30.9, 95: 33.8},
    },
    "Mujer": {
        "20-24": {5: 19.7, 10: 21.7, 20: 24.0, 30: 25.7, 40: 27.2, 50: 28.6, 60: 30.0, 70: 31.6, 80: 33.6, 90: 36.6, 95: 39.1},
        "25-29": {5: 20.0, 10: 22.0, 20: 24.5, 30: 26.3, 40: 27.9, 50: 29.4, 60: 30.9, 70: 32.6, 80: 34.6, 90: 37.4, 95: 39.7},
        "30-34": {5: 19.6, 10: 21.8, 20: 24.4, 30: 26.4, 40: 28.1, 50: 29.7, 60: 31.3, 70: 33.1, 80: 35.2, 90: 38.0, 95: 40.4},
        "35-39": {5: 19.0, 10: 21.3, 20: 24.1, 30: 26.2, 40: 28.0, 50: 29.7, 60: 31.4, 70: 33.2, 80: 35.4, 90: 38.4, 95: 40.8},
        "40-44": {5: 18.3, 10: 20.7, 20: 23.7, 30: 25.8, 40: 27.6, 50: 29.4, 60: 31.1, 70: 33.0, 80: 35.2, 90: 38.3, 95: 40.8},
        "45-49": {5: 17.6, 10: 20.1, 20: 23.1, 30: 25.2, 40: 27.1, 50: 28.9, 60: 30.6, 70: 32.5, 80: 34.8, 90: 37.9, 95: 40.4},
        "50-54": {5: 16.9, 10: 19.4, 20: 22.4, 30: 24.5, 40: 26.4, 50: 28.2, 60: 29.9, 70: 31.8, 80: 34.0, 90: 37.1, 95: 39.7},
        "55-59": {5: 16.1, 10: 18.5, 20: 21.5, 30: 23.7, 40: 25.5, 50: 27.3, 60: 29.0, 70: 30.9, 80: 33.0, 90: 36.1, 95: 38.6},
        "60-64": {5: 15.2, 10: 17.6, 20: 20.6, 30: 22.7, 40: 24.5, 50: 26.2, 60: 27.9, 70: 29.7, 80: 31.8, 90: 34.9, 95: 37.4},
        "65-69": {5: 14.3, 10: 16.6, 20: 19.5, 30: 21.6, 40: 23.3, 50: 25.0, 60: 26.6, 70: 28.4, 80: 30.5, 90: 33.4, 95: 35.8},
        "70-74": {5: 13.2, 10: 15.5, 20: 18.3, 30: 20.3, 40: 22.0, 50: 23.6, 60: 25.2, 70: 26.9, 80: 28.9, 90: 31.8, 95: 34.1},
        "75-79": {5: 12.0, 10: 14.3, 20: 17.0, 30: 18.9, 40: 20.5, 50: 22.1, 60: 23.6, 70: 25.2, 80: 27.2, 90: 29.9, 95: 32.2},
        "80-84": {5: 10.7, 10: 12.9, 20: 15.5, 30: 17.4, 40: 18.9, 50: 20.4, 60: 21.9, 70: 23.5, 80: 25.3, 90: 28.0, 95: 30.2},
        "85-89": {5: 9.3, 10: 11.4, 20: 13.9, 30: 15.7, 40: 17.2, 50: 18.6, 60: 20.0, 70: 21.5, 80: 23.3, 90: 25.9, 95: 28.0},
        "90-94": {5: 7.8, 10: 9.8, 20: 12.2, 30: 13.9, 40: 15.3, 50: 16.7, 60: 18.0, 70: 19.5, 80: 21.2, 90: 23.6, 95: 25.7},
        "95-99": {5: 6.1, 10: 8.0, 20: 10.3, 30: 11.9, 40: 13.3, 50: 14.6, 60: 15.9, 70: 17.3, 80: 18.9, 90: 21.2, 95: 23.2},
        "+100": {5: 4.2, 10: 6.1, 20: 8.3, 30: 9.8, 40: 11.2, 50: 12.4, 60: 13.6, 70: 14.9, 80: 16.5, 90: 18.7, 95: 20.6},
    }
}

TABLA_SILLA = {
    "Hombre": {
        "65-69": {10: 12, 20: 13, 30: 14, 40: 15, 50: 16, 60: 16, 70: 17, 80: 19, 90: 21, 100: 28},
        "70-74": {10: 11, 20: 13, 30: 14, 40: 15, 50: 15, 60: 16, 70: 17, 80: 18, 90: 20, 100: 29},
        "75-79": {10: 10, 20: 12, 30: 13, 40: 14, 50: 14, 60: 15, 70: 16, 80: 17, 90: 19, 100: 25},
        "80-84": {10: 9, 20: 10, 30: 11, 40: 12, 50: 14, 60: 15, 70: 16, 80: 17, 90: 18, 100: 22},
        "+84": {10: 9, 20: 9, 30: 12, 40: 13, 50: 14, 60: 14, 70: 16, 80: 18, 90: 20, 100: 21},
    },
    "Mujer": {
        "65-69": {10: 11, 20: 12, 30: 13, 40: 14, 50: 15, 60: 15, 70: 16, 80: 17, 90: 19, 100: 30},
        "70-74": {10: 10, 20: 12, 30: 12, 40: 13, 50: 14, 60: 15, 70: 16, 80: 17, 90: 19, 100: 27},
        "75-79": {10: 10, 20: 11, 30: 12, 40: 13, 50: 14, 60: 14, 70: 15, 80: 16, 90: 18, 100: 24},
        "80-84": {10: 9, 20: 10, 30: 11, 40: 12, 50: 13, 60: 14, 70: 15, 80: 16, 90: 18, 100: 24},
        "+84": {10: 6, 20: 8, 30: 9, 40: 11, 50: 12, 60: 14, 70: 14, 80: 16, 90: 17, 100: 18},
    }
}

# =========================================================
# BASE DE DATOS
# =========================================================
def guardar_evaluacion(paciente_nombre, sexo, edad, prueba, valor_medido, percentil, clasificacion):
    payload = {
        "fecha": datetime.now().strftime("%Y-%m-%d"),
        "paciente": str(paciente_nombre).strip(),
        "sexo": str(sexo).strip().lower(),
        "edad": int(edad),
        "prueba": str(prueba).strip(),
        "valor_medido": float(valor_medido),
        "percentil": round(float(percentil), 1) if percentil is not None else None,
        "clasificacion": str(clasificacion).strip()
    }
    return supabase.table("evaluaciones").insert(payload).execute()


def guardar_paciente(nombre, sexo, talla_m):
    nombre_limpio = str(nombre).strip()
    sexo_limpio = str(sexo).strip().lower()

    if not nombre_limpio:
        raise ValueError("El nombre del paciente está vacío.")

    if talla_m is None or float(talla_m) <= 0:
        raise ValueError("La talla debe ser mayor a 0.")

    respuesta = supabase.table("pacientes").select("id,nombre").execute()
    existentes = respuesta.data if respuesta.data else []

    for p in existentes:
        if str(p["nombre"]).strip().lower() == nombre_limpio.lower():
            raise ValueError("Ese paciente ya existe.")

    payload = {
        "nombre": nombre_limpio,
        "sexo": sexo_limpio,
        "talla_m": round(float(talla_m), 2)
    }
    return supabase.table("pacientes").insert(payload).execute()


def guardar_peso(paciente_id, fecha_medicion, peso_kg, talla_m):
    if paciente_id is None:
        raise ValueError("No se encontró el id del paciente.")

    if talla_m is None or float(talla_m) <= 0:
        raise ValueError("El paciente no tiene una talla válida cargada.")

    if peso_kg is None or float(peso_kg) <= 0:
        raise ValueError("El peso debe ser mayor a 0.")

    imc = round(float(peso_kg) / (float(talla_m) ** 2), 2)

    payload = {
        "paciente_id": int(paciente_id),
        "fecha": str(fecha_medicion),
        "peso_kg": float(peso_kg),
        "imc": imc
    }

    return supabase.table("seguimiento_peso").insert(payload).execute()


def clasificar_imc(imc_calculado):
    if imc_calculado < 18.5:
        return "Bajo peso", "🔵"
    elif imc_calculado < 25:
        return "Normal", "🟢"
    elif imc_calculado < 30:
        return "Sobrepeso", "🟡"
    else:
        return "Obesidad", "🔴"


def eliminar_evaluacion(id_registro):
    return supabase.table("evaluaciones").delete().eq("id", id_registro).execute()


def obtener_historial_paciente(paciente):
    try:
        respuesta = (
            supabase.table("evaluaciones")
            .select("*")
            .eq("paciente", str(paciente).strip())
            .order("fecha")
            .execute()
        )
        if respuesta.data:
            return pd.DataFrame(respuesta.data)
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Error al leer historial: {e}")
        return pd.DataFrame()


def obtener_historial_peso(paciente_id):
    try:
        respuesta = (
            supabase.table("seguimiento_peso")
            .select("id,paciente_id,fecha,peso_kg,imc,created_at")
            .eq("paciente_id", int(paciente_id))
            .order("fecha")
            .execute()
        )

        if respuesta.data:
            return pd.DataFrame(respuesta.data)

        return pd.DataFrame(columns=["fecha", "peso_kg", "imc"])
    except Exception as e:
        st.error(f"Error al leer historial de peso: {e}")
        return pd.DataFrame(columns=["fecha", "peso_kg", "imc"])


def obtener_pacientes():
    try:
        respuesta = (
            supabase.table("pacientes")
            .select("id,nombre,sexo,talla_m")
            .order("nombre")
            .execute()
        )
        return respuesta.data if respuesta.data else []
    except Exception as e:
        st.error(f"Error al leer pacientes: {e}")
        return []


def guardar_inbody(
    paciente_id,
    fecha_estudio,
    peso_kg,
    talla_m,
    grasa_corporal_pct,
    masa_muscular_kg,
    agua_corporal_pct,
    grasa_visceral,
    metabolismo_basal,
    observaciones
):
    imc_calculado = round(float(peso_kg) / (float(talla_m) ** 2), 2) if peso_kg and talla_m else None

    payload = {
        "paciente_id": int(paciente_id),
        "fecha": str(fecha_estudio),
        "peso_kg": float(peso_kg) if peso_kg is not None else None,
        "imc": float(imc_calculado) if imc_calculado is not None else None,
        "grasa_corporal_pct": float(grasa_corporal_pct) if grasa_corporal_pct is not None else None,
        "masa_muscular_kg": float(masa_muscular_kg) if masa_muscular_kg is not None else None,
        "agua_corporal_pct": float(agua_corporal_pct) if agua_corporal_pct is not None else None,
        "grasa_visceral": float(grasa_visceral) if grasa_visceral is not None else None,
        "metabolismo_basal": float(metabolismo_basal) if metabolismo_basal is not None else None,
        "observaciones": str(observaciones).strip() if observaciones else None
    }

    return supabase.table("inbody_registros").insert(payload).execute()


def obtener_historial_inbody(paciente_id):
    try:
        respuesta = (
            supabase.table("inbody_registros")
            .select("*")
            .eq("paciente_id", int(paciente_id))
            .order("fecha", desc=True)
            .execute()
        )

        if respuesta.data:
            return pd.DataFrame(respuesta.data)

        return pd.DataFrame()

    except Exception as e:
        st.error(f"Error al leer historial de composición corporal: {e}")
        return pd.DataFrame()


def obtener_ficha_paciente(paciente_nombre):
    try:
        paciente_nombre = str(paciente_nombre).strip()

        respuesta_paciente = (
            supabase.table("pacientes")
            .select("id,nombre,sexo,talla_m,created_at")
            .eq("nombre", paciente_nombre)
            .limit(1)
            .execute()
        )

        datos_paciente = respuesta_paciente.data[0] if respuesta_paciente.data else {}
        df_historial = obtener_historial_paciente(paciente_nombre)

        cantidad_evaluaciones = 0
        ultima_fecha = "-"
        ultima_clasificacion = "-"
        ultima_prueba = "-"

        if not df_historial.empty:
            if "fecha" in df_historial.columns:
                df_historial["fecha"] = pd.to_datetime(df_historial["fecha"], errors="coerce")
                df_historial = df_historial.sort_values("fecha", ascending=False)

            cantidad_evaluaciones = len(df_historial)

            if "fecha" in df_historial.columns and pd.notna(df_historial.iloc[0]["fecha"]):
                ultima_fecha = df_historial.iloc[0]["fecha"].strftime("%d-%m-%Y")

            if "clasificacion" in df_historial.columns and pd.notna(df_historial.iloc[0]["clasificacion"]):
                ultima_clasificacion = str(df_historial.iloc[0]["clasificacion"])

            if "prueba" in df_historial.columns and pd.notna(df_historial.iloc[0]["prueba"]):
                ultima_prueba = str(df_historial.iloc[0]["prueba"])

        return {
            "id": datos_paciente.get("id"),
            "nombre": datos_paciente.get("nombre", paciente_nombre),
            "sexo": datos_paciente.get("sexo", "-"),
            "talla_m": datos_paciente.get("talla_m"),
            "cantidad_evaluaciones": cantidad_evaluaciones,
            "ultima_fecha": ultima_fecha,
            "ultima_clasificacion": ultima_clasificacion,
            "ultima_prueba": ultima_prueba
        }

    except Exception as e:
        st.error(f"Error al cargar ficha del paciente: {e}")
        return {
            "id": None,
            "nombre": paciente_nombre,
            "sexo": "-",
            "talla_m": None,
            "cantidad_evaluaciones": 0,
            "ultima_fecha": "-",
            "ultima_clasificacion": "-",
            "ultima_prueba": "-"
        }


def preparar_df_exportacion(df):
    if df is None or df.empty:
        return pd.DataFrame()

    df_out = df.copy()

    for col in df_out.columns:
        if "fecha" in col.lower() or "created" in col.lower():
            try:
                df_out[col] = pd.to_datetime(df_out[col], errors="coerce")
                df_out[col] = df_out[col].dt.strftime("%Y-%m-%d")
            except Exception:
                pass

    return df_out

# =========================================================
# UTILIDADES CLÍNICAS - COMPOSICIÓN CORPORAL
# =========================================================
def clasificacion_grasa_corporal(sexo, grasa_pct):
    sexo = str(sexo).strip().lower()

    if grasa_pct is None or pd.isna(grasa_pct):
        return "Sin clasificar"

    if sexo == "hombre":
        if grasa_pct < 8:
            return "Muy bajo"
        elif grasa_pct < 19:
            return "Normal"
        elif grasa_pct <= 25:
            return "Alto"
        else:
            return "Obesidad"

    if sexo == "mujer":
        if grasa_pct < 21:
            return "Muy bajo"
        elif grasa_pct < 33:
            return "Normal"
        elif grasa_pct <= 39:
            return "Alto"
        else:
            return "Obesidad"

    return "Sin clasificar"


def clasificacion_agua_corporal(sexo, agua_pct):
    sexo = str(sexo).strip().lower()

    if agua_pct is None or pd.isna(agua_pct):
        return "Sin clasificar"

    if sexo == "hombre":
        if agua_pct < 50:
            return "Bajo"
        elif agua_pct <= 65:
            return "Normal"
        else:
            return "Alto"

    if sexo == "mujer":
        if agua_pct < 45:
            return "Bajo"
        elif agua_pct <= 60:
            return "Normal"
        else:
            return "Alto"

    return "Sin clasificar"


def clasificacion_grasa_visceral(grasa_visceral):
    if grasa_visceral is None or pd.isna(grasa_visceral):
        return "Sin clasificar"

    if grasa_visceral <= 9:
        return "Normal"
    elif grasa_visceral <= 14:
        return "Alto"
    else:
        return "Muy alto"


def calcular_masa_muscular_relativa_pct(peso_kg, masa_muscular_kg):
    if peso_kg is None or masa_muscular_kg is None:
        return None
    if pd.isna(peso_kg) or pd.isna(masa_muscular_kg):
        return None
    if float(peso_kg) <= 0:
        return None
    return round((float(masa_muscular_kg) / float(peso_kg)) * 100, 2)


def clasificacion_masa_muscular_relativa(sexo, musculo_relativo_pct):
    sexo = str(sexo).strip().lower()

    if musculo_relativo_pct is None or pd.isna(musculo_relativo_pct):
        return "Sin clasificar"

    if sexo == "hombre":
        if musculo_relativo_pct < 33:
            return "Bajo"
        elif musculo_relativo_pct <= 39:
            return "Normal"
        else:
            return "Alto"

    if sexo == "mujer":
        if musculo_relativo_pct < 24:
            return "Bajo"
        elif musculo_relativo_pct <= 30:
            return "Normal"
        else:
            return "Alto"

    return "Sin clasificar"


def color_estado_corporal(estado):
    mapa = {
        "Normal": ("#2e7d32", "#ffffff"),
        "Bajo peso": ("#1976d2", "#ffffff"),
        "Riesgo sarcopénico": ("#ef6c00", "#ffffff"),
        "Sobrepeso": ("#f9a825", "#1f1f1f"),
        "Sobrepeso muscular": ("#00897b", "#ffffff"),
        "Obesidad": ("#c62828", "#ffffff"),
        "Riesgo cardiometabólico": ("#ad1457", "#ffffff"),
        "Riesgo cardiometabólico moderado": ("#6a1b9a", "#ffffff"),
        "Sin clasificar": ("#757575", "#ffffff")
    }
    return mapa.get(estado, ("#757575", "#ffffff"))


def generar_recomendacion_corporal(estado, clasif_grasa, clasif_visceral, clasif_musculo):
    if estado in ["Obesidad", "Riesgo cardiometabólico", "Riesgo cardiometabólico moderado"]:
        return "Programa de reducción de grasa + ejercicio de fuerza + control cardiometabólico."
    if estado == "Riesgo sarcopénico":
        return "Priorizar ejercicio de fuerza, aumento de masa muscular y seguimiento funcional."
    if estado == "Sobrepeso muscular":
        return "Mantener masa muscular, controlar evolución y ajustar plan nutricional según objetivo."
    if estado == "Sobrepeso":
        return "Plan de control de peso con actividad física regular y seguimiento de composición corporal."
    if estado == "Bajo peso":
        return "Evaluar aporte nutricional y preservar o mejorar masa muscular."
    if estado == "Normal":
        if clasif_musculo == "Bajo":
            return "Estado general aceptable, pero conviene reforzar trabajo de fuerza."
        return "Mantener hábitos actuales y seguimiento periódico."
    return "Completar datos clínicos y repetir control."


def evaluar_perfil_morfofuncional(sexo, peso_kg, talla_m, grasa_pct, masa_muscular_kg, agua_pct, grasa_visceral):
    imc = round(float(peso_kg) / (float(talla_m) ** 2), 2) if peso_kg and talla_m else None
    clasif_imc = clasificar_imc(imc)[0] if imc is not None else "Sin clasificar"
    clasif_grasa = clasificacion_grasa_corporal(sexo, grasa_pct)
    clasif_agua = clasificacion_agua_corporal(sexo, agua_pct)
    clasif_visceral = clasificacion_grasa_visceral(grasa_visceral)
    musculo_rel_pct = calcular_masa_muscular_relativa_pct(peso_kg, masa_muscular_kg)
    clasif_musculo = clasificacion_masa_muscular_relativa(sexo, musculo_rel_pct)

    motivos = []

    if clasif_imc == "Bajo peso":
        motivos.append("IMC en rango de bajo peso")
    elif clasif_imc == "Sobrepeso":
        motivos.append("IMC en rango de sobrepeso")
    elif clasif_imc == "Obesidad":
        motivos.append("IMC en rango de obesidad")

    if clasif_grasa in ["Alto", "Obesidad"]:
        motivos.append("% grasa corporal elevada")

    if clasif_visceral in ["Alto", "Muy alto"]:
        motivos.append(f"grasa visceral {clasif_visceral.lower()}")

    if clasif_musculo == "Bajo":
        motivos.append("masa muscular relativa baja")
    elif clasif_musculo == "Normal":
        motivos.append("masa muscular relativa normal")
    elif clasif_musculo == "Alto":
        motivos.append("masa muscular relativa alta")

    estado = "Normal"

    if clasif_visceral == "Muy alto":
        estado = "Riesgo cardiometabólico"
    elif clasif_visceral == "Alto" and clasif_grasa in ["Alto", "Obesidad"]:
        estado = "Riesgo cardiometabólico moderado"
    elif clasif_imc == "Obesidad" or (clasif_imc in ["Sobrepeso", "Obesidad"] and clasif_grasa == "Obesidad"):
        estado = "Obesidad"
    elif clasif_imc == "Sobrepeso" and clasif_grasa in ["Alto", "Obesidad"] and clasif_musculo != "Alto":
        estado = "Sobrepeso"
    elif clasif_imc == "Sobrepeso" and clasif_musculo == "Alto" and clasif_grasa == "Normal":
        estado = "Sobrepeso muscular"
    elif clasif_imc == "Normal" and clasif_musculo == "Bajo":
        estado = "Riesgo sarcopénico"
    elif clasif_imc == "Bajo peso":
        estado = "Bajo peso"
    else:
        estado = "Normal"

    recomendacion = generar_recomendacion_corporal(
        estado=estado,
        clasif_grasa=clasif_grasa,
        clasif_visceral=clasif_visceral,
        clasif_musculo=clasif_musculo
    )

    return {
        "imc": imc,
        "clasif_imc": clasif_imc,
        "clasif_grasa": clasif_grasa,
        "clasif_agua": clasif_agua,
        "clasif_visceral": clasif_visceral,
        "musculo_rel_pct": musculo_rel_pct,
        "clasif_musculo": clasif_musculo,
        "estado": estado,
        "motivos": motivos,
        "recomendacion": recomendacion
    }


def enriquecer_historial_corporal(df, sexo, talla_m):
    if df is None or df.empty:
        return pd.DataFrame()

    df = df.copy()

    for col in [
        "peso_kg",
        "imc",
        "grasa_corporal_pct",
        "masa_muscular_kg",
        "agua_corporal_pct",
        "grasa_visceral",
        "metabolismo_basal"
    ]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    diagnosticos = []
    sugerencias = []
    motivos_lista = []
    clasif_imc_lista = []
    clasif_grasa_lista = []
    clasif_agua_lista = []
    clasif_visceral_lista = []
    musculo_rel_lista = []
    clasif_musculo_lista = []

    for _, row in df.iterrows():
        res = evaluar_perfil_morfofuncional(
            sexo=sexo,
            peso_kg=row.get("peso_kg"),
            talla_m=talla_m,
            grasa_pct=row.get("grasa_corporal_pct"),
            masa_muscular_kg=row.get("masa_muscular_kg"),
            agua_pct=row.get("agua_corporal_pct"),
            grasa_visceral=row.get("grasa_visceral")
        )

        diagnosticos.append(res["estado"])
        sugerencias.append(res["recomendacion"])
        motivos_lista.append(" | ".join(res["motivos"]) if res["motivos"] else "")
        clasif_imc_lista.append(res["clasif_imc"])
        clasif_grasa_lista.append(res["clasif_grasa"])
        clasif_agua_lista.append(res["clasif_agua"])
        clasif_visceral_lista.append(res["clasif_visceral"])
        musculo_rel_lista.append(res["musculo_rel_pct"])
        clasif_musculo_lista.append(res["clasif_musculo"])

    df["diagnostico_corporal"] = diagnosticos
    df["sugerencia_corporal"] = sugerencias
    df["motivos_corporal"] = motivos_lista
    df["clasif_imc"] = clasif_imc_lista
    df["clasif_grasa"] = clasif_grasa_lista
    df["clasif_agua"] = clasif_agua_lista
    df["clasif_visceral"] = clasif_visceral_lista
    df["musculo_rel_pct"] = musculo_rel_lista
    df["clasif_musculo"] = clasif_musculo_lista

    return df


def agregar_identificacion_paciente(df, ficha):
    if df is None or df.empty:
        return pd.DataFrame()

    df_out = df.copy()
    df_out.insert(0, "PacienteID", ficha.get("id"))
    df_out.insert(0, "Talla_m", ficha.get("talla_m"))
    df_out.insert(0, "Sexo", ficha.get("sexo"))
    df_out.insert(0, "Paciente", ficha.get("nombre"))
    return df_out


def generar_df_analisis_cientifico(ficha, df_peso, df_inbody, df_eval):
    filas = []

    nombre = ficha.get("nombre")
    sexo = ficha.get("sexo")
    talla_m = ficha.get("talla_m")
    paciente_id = ficha.get("id")

    if df_peso is not None and not df_peso.empty:
        df_p = df_peso.copy()
        if "fecha" in df_p.columns:
            df_p["fecha"] = pd.to_datetime(df_p["fecha"], errors="coerce")

        for _, row in df_p.iterrows():
            filas.append({
                "PacienteID": paciente_id,
                "Paciente": nombre,
                "Sexo": sexo,
                "Talla_m": talla_m,
                "Fecha": row.get("fecha"),
                "TipoRegistro": "Peso_IMC",
                "Prueba": None,
                "ValorMedido": None,
                "Percentil": None,
                "ClasificacionFuncional": None,
                "Peso_kg": row.get("peso_kg"),
                "IMC": row.get("imc"),
                "GrasaCorporal_pct": None,
                "MasaMuscular_kg": None,
                "AguaCorporal_pct": None,
                "GrasaVisceral": None,
                "MetabolismoBasal": None,
                "MusculoRelativo_pct": None,
                "Clasif_IMC": clasificar_imc(row.get("imc"))[0] if pd.notna(row.get("imc")) else None,
                "Clasif_Grasa": None,
                "Clasif_Agua": None,
                "Clasif_Visceral": None,
                "Clasif_Musculo": None,
                "DiagnosticoCorporal": None,
                "SugerenciaCorporal": None,
                "MotivosCorporal": None,
                "Observaciones": None
            })

    if df_inbody is not None and not df_inbody.empty:
        df_c = enriquecer_historial_corporal(df_inbody, str(sexo).strip().lower(), talla_m)
        if "fecha" in df_c.columns:
            df_c["fecha"] = pd.to_datetime(df_c["fecha"], errors="coerce")

        for _, row in df_c.iterrows():
            filas.append({
                "PacienteID": paciente_id,
                "Paciente": nombre,
                "Sexo": sexo,
                "Talla_m": talla_m,
                "Fecha": row.get("fecha"),
                "TipoRegistro": "Composicion_Corporal",
                "Prueba": None,
                "ValorMedido": None,
                "Percentil": None,
                "ClasificacionFuncional": None,
                "Peso_kg": row.get("peso_kg"),
                "IMC": row.get("imc"),
                "GrasaCorporal_pct": row.get("grasa_corporal_pct"),
                "MasaMuscular_kg": row.get("masa_muscular_kg"),
                "AguaCorporal_pct": row.get("agua_corporal_pct"),
                "GrasaVisceral": row.get("grasa_visceral"),
                "MetabolismoBasal": row.get("metabolismo_basal"),
                "MusculoRelativo_pct": row.get("musculo_rel_pct"),
                "Clasif_IMC": row.get("clasif_imc"),
                "Clasif_Grasa": row.get("clasif_grasa"),
                "Clasif_Agua": row.get("clasif_agua"),
                "Clasif_Visceral": row.get("clasif_visceral"),
                "Clasif_Musculo": row.get("clasif_musculo"),
                "DiagnosticoCorporal": row.get("diagnostico_corporal"),
                "SugerenciaCorporal": row.get("sugerencia_corporal"),
                "MotivosCorporal": row.get("motivos_corporal"),
                "Observaciones": row.get("observaciones")
            })

    if df_eval is not None and not df_eval.empty:
        df_f = df_eval.copy()
        if "fecha" in df_f.columns:
            df_f["fecha"] = pd.to_datetime(df_f["fecha"], errors="coerce")

        for _, row in df_f.iterrows():
            filas.append({
                "PacienteID": paciente_id,
                "Paciente": nombre,
                "Sexo": sexo,
                "Talla_m": talla_m,
                "Fecha": row.get("fecha"),
                "TipoRegistro": "Evaluacion_Funcional",
                "Prueba": row.get("prueba"),
                "ValorMedido": row.get("valor_medido"),
                "Percentil": row.get("percentil"),
                "ClasificacionFuncional": row.get("clasificacion"),
                "Peso_kg": None,
                "IMC": None,
                "GrasaCorporal_pct": None,
                "MasaMuscular_kg": None,
                "AguaCorporal_pct": None,
                "GrasaVisceral": None,
                "MetabolismoBasal": None,
                "MusculoRelativo_pct": None,
                "Clasif_IMC": None,
                "Clasif_Grasa": None,
                "Clasif_Agua": None,
                "Clasif_Visceral": None,
                "Clasif_Musculo": None,
                "DiagnosticoCorporal": None,
                "SugerenciaCorporal": None,
                "MotivosCorporal": None,
                "Observaciones": None
            })

    df_final = pd.DataFrame(filas)

    if not df_final.empty:
        df_final["Fecha"] = pd.to_datetime(df_final["Fecha"], errors="coerce")
        df_final = df_final.sort_values(["Paciente", "Fecha", "TipoRegistro"], ascending=[True, True, True]).reset_index(drop=True)

    return df_final

# =========================================================
# EXPORTACIÓN
# =========================================================
def generar_excel_paciente(ficha, df_peso, df_inbody, df_eval):
    output = BytesIO()

    ficha_df = pd.DataFrame([{
        "PacienteID": ficha["id"],
        "Paciente": ficha["nombre"],
        "Sexo": ficha["sexo"],
        "Talla_m": ficha["talla_m"],
        "CantidadEvaluaciones": ficha["cantidad_evaluaciones"],
        "UltimaFechaEvaluacion": ficha["ultima_fecha"],
        "UltimaClasificacion": ficha["ultima_clasificacion"],
        "UltimaPrueba": ficha["ultima_prueba"]
    }])

    df_peso_export = agregar_identificacion_paciente(df_peso, ficha)
    df_inbody_enriquecido = enriquecer_historial_corporal(
        df_inbody,
        str(ficha["sexo"]).strip().lower(),
        ficha["talla_m"]
    )
    df_inbody_export = agregar_identificacion_paciente(df_inbody_enriquecido, ficha)
    df_eval_export = agregar_identificacion_paciente(df_eval, ficha)

    df_analisis = generar_df_analisis_cientifico(
        ficha=ficha,
        df_peso=df_peso,
        df_inbody=df_inbody,
        df_eval=df_eval
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        preparar_df_exportacion(ficha_df).to_excel(writer, sheet_name="Ficha", index=False)
        preparar_df_exportacion(df_peso_export).to_excel(writer, sheet_name="Peso_IMC", index=False)
        preparar_df_exportacion(df_inbody_export).to_excel(writer, sheet_name="Composicion_Corporal", index=False)
        preparar_df_exportacion(df_eval_export).to_excel(writer, sheet_name="Evaluaciones", index=False)
        preparar_df_exportacion(df_analisis).to_excel(writer, sheet_name="Analisis_Cientifico", index=False)

    output.seek(0)
    return output


def escribir_bloque_pdf(pdf, y, titulo, df, columnas, anchos):
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(40, y, titulo)
    y -= 18

    pdf.setFont("Helvetica-Bold", 8)
    x = 40
    for col, ancho in zip(columnas, anchos):
        pdf.drawString(x, y, str(col))
        x += ancho

    y -= 14
    pdf.setFont("Helvetica", 8)

    if df is None or df.empty:
        pdf.drawString(40, y, "Sin datos")
        return y - 20

    for _, row in df.iterrows():
        x = 40
        for col, ancho in zip(columnas, anchos):
            valor = row.get(col, "")
            if pd.isna(valor):
                valor = ""
            valor = str(valor)
            if len(valor) > 24:
                valor = valor[:21] + "..."
            pdf.drawString(x, y, valor)
            x += ancho

        y -= 12

        if y < 70:
            pdf.showPage()
            y = 750
            pdf.setFont("Helvetica", 8)

    return y - 12


def generar_pdf_paciente(ficha, df_peso, df_inbody, df_eval):
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    y = 760

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(40, y, "Reporte completo del paciente")
    y -= 24

    pdf.setFont("Helvetica", 11)
    pdf.drawString(40, y, f"Paciente: {ficha['nombre']}")
    y -= 16
    pdf.drawString(40, y, f"Sexo: {ficha['sexo']}")
    y -= 16
    pdf.drawString(40, y, f"Talla: {ficha['talla_m']} m")
    y -= 16
    pdf.drawString(40, y, f"Fecha del reporte: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    y -= 24

    df_peso_pdf = preparar_df_exportacion(df_peso)
    df_inbody_pdf = preparar_df_exportacion(df_inbody)
    df_inbody_pdf = enriquecer_historial_corporal(
        df_inbody_pdf,
        str(ficha["sexo"]).strip().lower(),
        ficha["talla_m"]
    )
    df_eval_pdf = preparar_df_exportacion(df_eval)

    if df_inbody_pdf is not None and not df_inbody_pdf.empty:
        columnas_requeridas = [
            "diagnostico_corporal",
            "sugerencia_corporal",
            "motivos_corporal"
        ]
        for c in columnas_requeridas:
            if c not in df_inbody_pdf.columns:
                df_inbody_pdf[c] = "-"

        ultimo_corporal = df_inbody_pdf.copy()

        if "fecha" in ultimo_corporal.columns:
            ultimo_corporal["fecha"] = pd.to_datetime(ultimo_corporal["fecha"], errors="coerce")
            ultimo_corporal = ultimo_corporal.sort_values("fecha", ascending=False)

        ult = ultimo_corporal.iloc[0]

        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(40, y, "Diagnóstico corporal actual")
        y -= 18

        pdf.setFont("Helvetica", 10)
        pdf.drawString(40, y, f"Diagnóstico: {str(ult.get('diagnostico_corporal', '-'))}")
        y -= 14
        pdf.drawString(40, y, f"Sugerencia: {str(ult.get('sugerencia_corporal', '-'))[:110]}")
        y -= 14
        pdf.drawString(40, y, f"Motivos: {str(ult.get('motivos_corporal', '-'))[:110]}")
        y -= 22

    y = escribir_bloque_pdf(
        pdf,
        y,
        "Historial de peso e IMC",
        df_peso_pdf,
        ["fecha", "peso_kg", "imc"],
        [120, 120, 120]
    )

    y = escribir_bloque_pdf(
        pdf,
        y,
        "Historial de composición corporal",
        df_inbody_pdf,
        [
            "fecha",
            "peso_kg",
            "imc",
            "grasa_corporal_pct",
            "masa_muscular_kg",
            "diagnostico_corporal",
            "motivos_corporal"
        ],
        [60, 50, 45, 60, 60, 90, 120]
    )

    y = escribir_bloque_pdf(
        pdf,
        y,
        "Evaluaciones funcionales",
        df_eval_pdf,
        ["fecha", "prueba", "valor_medido", "percentil", "clasificacion"],
        [70, 150, 70, 70, 100]
    )

    pdf.save()
    buffer.seek(0)
    return buffer

# =========================================================
# UTILIDADES CLÍNICAS - FUNCIONALES
# =========================================================
def clasificar_percentil(percentil):
    if percentil is None:
        return "Sin clasificar"
    if percentil < 10:
        return "Muy bajo"
    if percentil < 25:
        return "Bajo"
    if percentil < 50:
        return "Ligeramente bajo"
    if percentil < 75:
        return "Normal"
    if percentil < 90:
        return "Bueno"
    return "Muy bueno"


def color_clasificacion(clasificacion):
    colores = {
        "Muy bajo": "#d32f2f",
        "Bajo": "#f57c00",
        "Ligeramente bajo": "#fbc02d",
        "Normal": "#388e3c",
        "Bueno": "#1976d2",
        "Muy bueno": "#00796b",
        "Sin clasificar": "#757575"
    }
    return colores.get(clasificacion, "#757575")


def rango_percentilar(percentil):
    if percentil is None:
        return "Sin rango"
    if percentil < 3:
        return "Menor a P3"
    if percentil < 10:
        return "Entre P3 y P10"
    if percentil < 25:
        return "Entre P10 y P25"
    if percentil < 50:
        return "Entre P25 y P50"
    if percentil < 75:
        return "Entre P50 y P75"
    if percentil < 90:
        return "Entre P75 y P90"
    if percentil < 97:
        return "Entre P90 y P97"
    return "Mayor a P97"


def interpretacion_clinica(clasificacion):
    textos = {
        "Muy bajo": "Resultado marcadamente por debajo del rango funcional esperado.",
        "Bajo": "Resultado por debajo del rango funcional esperado.",
        "Ligeramente bajo": "Resultado levemente inferior al rango esperado.",
        "Normal": "Resultado dentro del rango funcional esperado.",
        "Bueno": "Resultado superior al promedio esperado.",
        "Muy bueno": "Resultado claramente superior al rango esperado."
    }
    return textos.get(clasificacion, "Sin interpretación disponible.")


def interpolar_percentil(valor_medido, tabla_percentiles):
    puntos = sorted(tabla_percentiles.items(), key=lambda x: x[1])

    if valor_medido < puntos[0][1]:
        return float(puntos[0][0])

    if valor_medido > puntos[-1][1]:
        return float(puntos[-1][0])

    for i in range(len(puntos) - 1):
        p1, v1 = puntos[i]
        p2, v2 = puntos[i + 1]

        if v1 <= valor_medido <= v2:
            if v2 == v1:
                return float(p1)
            return float(p1 + (valor_medido - v1) * (p2 - p1) / (v2 - v1))

    return None


def grupo_edad_prension(edad):
    edad = int(edad)
    if edad >= 100:
        return "+100"
    inicio = max(20, min(95, 20 + ((edad - 20) // 5) * 5))
    fin = inicio + 4
    return f"{inicio}-{fin}"


def grupo_edad_silla(edad):
    edad = int(edad)
    if edad >= 85:
        return "+84"
    if 65 <= edad <= 69:
        return "65-69"
    if 70 <= edad <= 74:
        return "70-74"
    if 75 <= edad <= 79:
        return "75-79"
    if 80 <= edad <= 84:
        return "80-84"
    return None


def calcular_resultado(prueba, sexo, edad, altura, valor_medido):
    sexo = str(sexo).strip()
    edad = int(edad)
    valor_medido = float(valor_medido)

    if prueba == "Caminata 6 minutos":
        altura = int(altura)
        altura_ref = min(TABLA_CAMINATA_6M.keys(), key=lambda x: abs(x - altura))
        edad_ref = min(TABLA_CAMINATA_6M[altura_ref].keys(), key=lambda x: abs(x - edad))

        percentiles = TABLA_CAMINATA_6M[altura_ref][edad_ref]
        percentil_estimado = interpolar_percentil(valor_medido, percentiles)

        if percentil_estimado is None:
            return None, "Sin clasificar", "-", "-", "-"

        percentil_estimado = round(percentil_estimado, 1)
        clasificacion = clasificar_percentil(percentil_estimado)
        referencia_p50 = percentiles[50]

        return (
            percentil_estimado,
            clasificacion,
            f"{referencia_p50} m",
            f"Altura ref.: {altura_ref} cm",
            f"Edad ref.: {edad_ref} años"
        )

    if prueba == "Prensión manual":
        grupo = grupo_edad_prension(edad)
        percentiles = TABLA_PRENSION[sexo][grupo]
        percentil_estimado = interpolar_percentil(valor_medido, percentiles)

        if percentil_estimado is None:
            return None, "Sin clasificar", "-", "-", "-"

        percentil_estimado = round(percentil_estimado, 1)
        clasificacion = clasificar_percentil(percentil_estimado)
        referencia_p50 = percentiles[50]

        return (
            percentil_estimado,
            clasificacion,
            f"{referencia_p50} kg",
            "-",
            f"Grupo etario: {grupo}"
        )

    if prueba == "Levantarse de la silla":
        grupo = grupo_edad_silla(edad)

        if grupo is None:
            return None, "Sin clasificar", "-", "-", "Edad fuera del rango de la tabla (65+ años)"

        percentiles = TABLA_SILLA[sexo][grupo]
        percentil_estimado = interpolar_percentil(valor_medido, percentiles)

        if percentil_estimado is None:
            return None, "Sin clasificar", "-", "-", "-"

        percentil_estimado = round(percentil_estimado, 1)
        clasificacion = clasificar_percentil(percentil_estimado)
        referencia_p50 = percentiles[50]

        return (
            percentil_estimado,
            clasificacion,
            f"{referencia_p50} rep",
            "-",
            f"Grupo etario: {grupo}"
        )

    return None, "Sin clasificar", "-", "-", "-"

# =========================================================
# UI
# =========================================================
st.title("Calculadora de Condición Física")

with st.expander("➕ Nuevo paciente"):
    nuevo_nombre = st.text_input("Nombre del nuevo paciente", key="nuevo_nombre_alta")
    nuevo_sexo = st.selectbox("Sexo del nuevo paciente", ["hombre", "mujer"], key="nuevo_sexo_alta")
    nueva_talla = st.number_input(
        "Talla (m)",
        min_value=0.50,
        max_value=2.50,
        value=1.70,
        step=0.01,
        format="%.2f",
        key="nueva_talla_alta"
    )

    if st.button("Guardar paciente", key="btn_guardar_paciente"):
        if not nuevo_nombre.strip():
            st.warning("Ingresá el nombre del paciente.")
        else:
            try:
                guardar_paciente(nuevo_nombre, nuevo_sexo, nueva_talla)
                st.success("Paciente agregado correctamente.")
                st.rerun()
            except Exception as e:
                st.error(f"Error al guardar paciente: {e}")

pacientes = obtener_pacientes()
opciones_pacientes = [p["nombre"] for p in pacientes]

if not opciones_pacientes:
    st.warning("No hay pacientes cargados. Agregá uno desde '➕ Nuevo paciente'.")
    st.stop()

# =========================================================
# ENCABEZADO
# =========================================================
top1, top2, top3 = st.columns([2, 1, 1])

with top1:
    paciente_nombre = st.selectbox(
        "Seleccionar paciente",
        opciones_pacientes,
        key="selector_paciente"
    )

paciente_actual = next((p for p in pacientes if p["nombre"] == paciente_nombre), None)
paciente_id = paciente_actual["id"] if paciente_actual else None

# =========================================================
# FICHA + DATAFRAMES BASE
# =========================================================
ficha = obtener_ficha_paciente(paciente_nombre)
df_peso_export = obtener_historial_peso(paciente_id) if paciente_id is not None else pd.DataFrame()
df_inbody_export = obtener_historial_inbody(paciente_id) if paciente_id is not None else pd.DataFrame()
df_eval_export = obtener_historial_paciente(paciente_nombre) if paciente_nombre else pd.DataFrame()

excel_buffer = generar_excel_paciente(ficha, df_peso_export, df_inbody_export, df_eval_export)
pdf_buffer = generar_pdf_paciente(ficha, df_peso_export, df_inbody_export, df_eval_export)

with top2:
    st.download_button(
        label="Descargar Excel",
        data=excel_buffer.getvalue(),
        file_name=f"{paciente_nombre.replace(' ', '_')}_reporte.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"descargar_excel_{paciente_id}"
    )

with top3:
    st.download_button(
        label="Descargar PDF",
        data=pdf_buffer.getvalue(),
        file_name=f"{paciente_nombre.replace(' ', '_')}_reporte.pdf",
        mime="application/pdf",
        key=f"descargar_pdf_{paciente_id}"
    )

st.divider()

# =========================================================
# TARJETAS RESUMEN
# =========================================================
k1, k2, k3, k4 = st.columns(4)

with k1:
    with st.container(border=True):
        st.markdown("#### Paciente")
        st.write(f"**Nombre:** {ficha['nombre']}")
        st.write(f"**Sexo:** {str(ficha['sexo']).capitalize() if ficha['sexo'] != '-' else '-'}")
        if ficha["talla_m"] is not None:
            st.write(f"**Talla:** {float(ficha['talla_m']):.2f} m")
        else:
            st.write("**Talla:** -")

with k2:
    with st.container(border=True):
        st.markdown("#### Peso / IMC")
        if not df_peso_export.empty:
            df_peso_tmp = df_peso_export.copy()
            df_peso_tmp["fecha"] = pd.to_datetime(df_peso_tmp["fecha"], errors="coerce")
            df_peso_tmp = df_peso_tmp.dropna(subset=["fecha"]).sort_values("fecha", ascending=False)
            ultimo_peso = df_peso_tmp.iloc[0]
            st.write(f"**Peso:** {float(ultimo_peso['peso_kg']):.1f} kg")
            st.write(f"**IMC:** {float(ultimo_peso['imc']):.2f}")
            st.write(f"**Fecha:** {ultimo_peso['fecha'].strftime('%d-%m-%Y')}")
        else:
            st.write("Sin registros")

with k3:
    with st.container(border=True):
        st.markdown("#### Estado corporal")
        if not df_inbody_export.empty and ficha["talla_m"] is not None:
            df_corporal_tmp = enriquecer_historial_corporal(
                df_inbody_export,
                str(ficha["sexo"]).strip().lower(),
                ficha["talla_m"]
            )
            df_corporal_tmp["fecha"] = pd.to_datetime(df_corporal_tmp["fecha"], errors="coerce")
            df_corporal_tmp = df_corporal_tmp.dropna(subset=["fecha"]).sort_values("fecha", ascending=False)
            ultimo_corporal = df_corporal_tmp.iloc[0]

            st.write(f"**Estado:** {ultimo_corporal.get('diagnostico_corporal', '-')}")
            st.write(f"**Grasa:** {ultimo_corporal.get('clasif_grasa', '-')}")
            st.write(f"**Músculo:** {ultimo_corporal.get('clasif_musculo', '-')}")
        else:
            st.write("Sin registros")

with k4:
    with st.container(border=True):
        st.markdown("#### Última evaluación")
        st.write(f"**Fecha:** {ficha['ultima_fecha']}")
        st.write(f"**Prueba:** {ficha['ultima_prueba']}")
        st.write(f"**Clasificación:** {ficha['ultima_clasificacion']}")

st.divider()

# =========================================================
# ZONA PRINCIPAL DE TRABAJO
# =========================================================
left, right = st.columns([1, 1])

with left:
    st.markdown("## Peso e IMC")

    with st.container(border=True):
        if ficha["talla_m"] is None or float(ficha["talla_m"]) <= 0:
            st.warning("Este paciente no tiene una talla válida cargada en la tabla pacientes.")
        else:
            col_p1, col_p2, col_p3 = st.columns(3)

            with col_p1:
                fecha_peso = st.date_input(
                    "Fecha de peso",
                    value=date.today(),
                    key=f"fecha_peso_{paciente_id}"
                )

            with col_p2:
                peso_kg = st.number_input(
                    "Peso (kg)",
                    min_value=0.0,
                    max_value=300.0,
                    value=70.0,
                    step=0.1,
                    format="%.1f",
                    key=f"peso_kg_{paciente_id}"
                )

            with col_p3:
                imc_calculado = round(float(peso_kg) / (float(ficha["talla_m"]) ** 2), 2)
                clasificacion_imc, color_imc = clasificar_imc(imc_calculado)
                st.markdown(f"**IMC:** {imc_calculado:.2f}")
                st.markdown(f"**Clasificación:** {color_imc} {clasificacion_imc}")

            if st.button("Guardar peso", key=f"btn_guardar_peso_{paciente_id}"):
                try:
                    guardar_peso(
                        paciente_id=paciente_id,
                        fecha_medicion=fecha_peso,
                        peso_kg=peso_kg,
                        talla_m=ficha["talla_m"]
                    )
                    st.success("Peso guardado correctamente.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al guardar peso: {e}")

    st.markdown("## Composición corporal")

    with st.container(border=True):
        if ficha["talla_m"] is None or float(ficha["talla_m"]) <= 0:
            st.warning("Para cargar composición corporal primero hay que tener una talla válida en el paciente.")
        else:
            sexo_corporal = str(ficha["sexo"]).strip().lower()

            col_c1, col_c2 = st.columns(2)

            with col_c1:
                fecha_inbody = st.date_input("Fecha estudio", value=date.today(), key=f"inbody_fecha_{paciente_id}")
                peso_inbody = st.number_input("Peso (kg)", min_value=0.0, max_value=300.0, step=0.1, key=f"inbody_peso_{paciente_id}")
                imc_inbody_calc = round(float(peso_inbody) / (float(ficha["talla_m"]) ** 2), 2) if peso_inbody and ficha["talla_m"] else 0.0
                st.markdown(f"**IMC calculado:** {imc_inbody_calc:.2f}")
                grasa_pct = st.number_input("% grasa corporal", min_value=0.0, max_value=80.0, step=0.1, key=f"inbody_grasa_{paciente_id}")

            with col_c2:
                masa_muscular = st.number_input("Masa muscular (kg)", min_value=0.0, max_value=100.0, step=0.1, key=f"inbody_musculo_{paciente_id}")
                agua_pct = st.number_input("% agua corporal", min_value=0.0, max_value=100.0, step=0.1, key=f"inbody_agua_{paciente_id}")
                grasa_visceral = st.number_input("Grasa visceral", min_value=0.0, max_value=30.0, step=0.1, key=f"inbody_visceral_{paciente_id}")
                metabolismo = st.number_input("Metabolismo basal", min_value=0.0, max_value=4000.0, step=10.0, key=f"inbody_metabolismo_{paciente_id}")

            observaciones_inbody = st.text_area("Observaciones", key=f"inbody_obs_{paciente_id}")

            resultado_corporal = evaluar_perfil_morfofuncional(
                sexo=sexo_corporal,
                peso_kg=peso_inbody,
                talla_m=ficha["talla_m"],
                grasa_pct=grasa_pct,
                masa_muscular_kg=masa_muscular,
                agua_pct=agua_pct,
                grasa_visceral=grasa_visceral
            )

            bg_estado, fg_estado = color_estado_corporal(resultado_corporal["estado"])

            st.markdown(
                f"""
                <div class="result-card" style="background-color:{bg_estado}; color:{fg_estado};">
                    Diagnóstico corporal: {resultado_corporal["estado"]}
                </div>
                """,
                unsafe_allow_html=True
            )

            col_r1, col_r2, col_r3 = st.columns(3)
            with col_r1:
                st.write(f"**IMC:** {resultado_corporal['imc'] if resultado_corporal['imc'] is not None else '-'}")
                st.write(f"**Clasificación IMC:** {resultado_corporal['clasif_imc']}")
                st.write(f"**% grasa:** {resultado_corporal['clasif_grasa']}")

            with col_r2:
                st.write(f"**% agua corporal:** {resultado_corporal['clasif_agua']}")
                st.write(f"**Grasa visceral:** {resultado_corporal['clasif_visceral']}")
                st.write(f"**Músculo relativo %:** {resultado_corporal['musculo_rel_pct'] if resultado_corporal['musculo_rel_pct'] is not None else '-'}")

            with col_r3:
                st.write(f"**Clasificación muscular:** {resultado_corporal['clasif_musculo']}")
                st.write(f"**Metabolismo basal:** {metabolismo if metabolismo is not None else '-'}")
                st.write(f"**Sexo de referencia:** {sexo_corporal.capitalize() if sexo_corporal else '-'}")

            motivos_texto = resultado_corporal["motivos"] if resultado_corporal["motivos"] else ["Sin hallazgos relevantes"]
            motivos_html = "".join([f"<li>{m}</li>" for m in motivos_texto])

            st.markdown(
                f"""
                <div class="motivo-box">
                    <b>Motivos:</b>
                    <ul style="margin-top:8px; margin-bottom:0;">
                        {motivos_html}
                    </ul>
                </div>
                """,
                unsafe_allow_html=True
            )

            st.markdown(
                f"""
                <div class="reco-box">
                    <b>Sugerencia:</b><br>
                    {resultado_corporal["recomendacion"]}
                </div>
                """,
                unsafe_allow_html=True
            )

            if st.button("Guardar composición corporal", key=f"guardar_inbody_{paciente_id}"):
                try:
                    guardar_inbody(
                        paciente_id=paciente_id,
                        fecha_estudio=fecha_inbody,
                        peso_kg=peso_inbody,
                        talla_m=ficha["talla_m"],
                        grasa_corporal_pct=grasa_pct,
                        masa_muscular_kg=masa_muscular,
                        agua_corporal_pct=agua_pct,
                        grasa_visceral=grasa_visceral,
                        metabolismo_basal=metabolismo,
                        observaciones=observaciones_inbody
                    )
                    st.success("Composición corporal guardada correctamente")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al guardar composición corporal: {e}")

with right:
    st.markdown("## Evaluación funcional")

    with st.container(border=True):
        paciente_sexo_guardado = next((p["sexo"] for p in pacientes if p["nombre"] == paciente_nombre), None)

        prueba = st.selectbox(
            "Seleccionar prueba",
            ["Caminata 6 minutos", "Prensión manual", "Levantarse de la silla"],
            key="selector_prueba"
        )

        sexo_default = "Hombre"
        if paciente_sexo_guardado == "mujer":
            sexo_default = "Mujer"
        elif paciente_sexo_guardado == "hombre":
            sexo_default = "Hombre"

        sexo = st.selectbox(
            "Sexo",
            ["Hombre", "Mujer"],
            index=0 if sexo_default == "Hombre" else 1,
            key="selector_sexo"
        )

        altura = None
        valor_medido = None

        if prueba == "Caminata 6 minutos":
            edad = st.selectbox("Edad", list(range(40, 81)), index=20, key="edad_caminata")
            altura = st.selectbox("Altura (cm)", [150, 160, 170, 180, 190], index=2, key="altura_caminata")
            valor_medido = st.number_input(
                "Distancia caminada (metros)",
                min_value=0.0,
                max_value=2000.0,
                value=600.0,
                step=1.0,
                format="%.2f",
                key="valor_caminata"
            )

        elif prueba == "Prensión manual":
            edad = st.selectbox("Edad", list(range(20, 101)), index=45, key="edad_prension")
            valor_medido = st.number_input(
                "Fuerza de prensión (kg)",
                min_value=0.0,
                max_value=100.0,
                value=25.0,
                step=0.1,
                format="%.1f",
                key="valor_prension"
            )

        elif prueba == "Levantarse de la silla":
            edad = st.selectbox("Edad", list(range(65, 101)), index=10, key="edad_silla")
            valor_medido = st.number_input(
                "Cantidad de repeticiones",
                min_value=0.0,
                max_value=60.0,
                value=12.0,
                step=1.0,
                format="%.0f",
                key="valor_silla"
            )

        percentil, clasificacion, referencia_p50, referencia_altura, referencia_edad = calcular_resultado(
            prueba=prueba,
            sexo=sexo,
            edad=edad,
            altura=altura,
            valor_medido=valor_medido
        )

        color = color_clasificacion(clasificacion)

        st.markdown(
            f"""
            <div style="
                background-color:{color};
                color:white;
                padding:10px 12px;
                border-radius:8px;
                text-align:center;
                font-size:18px;
                font-weight:600;
                margin-top:18px;
                margin-bottom:10px;
            ">
                {clasificacion}
            </div>
            """,
            unsafe_allow_html=True
        )

        st.markdown(
            f"""
            <div style="
                background-color:#dff0e6;
                color:#1b5e20;
                padding:8px 12px;
                border-radius:8px;
                font-size:15px;
                margin-bottom:14px;
            ">
                Percentil estimado: <b>P{percentil if percentil is not None else "-"}</b>
            </div>
            """,
            unsafe_allow_html=True
        )

        st.write(f"**Rango percentilar:** {rango_percentilar(percentil)}")
        st.write(f"**Referencia P50:** {referencia_p50}")

        if referencia_altura != "-":
            st.write(f"**Referencia de altura:** {referencia_altura}")

        if referencia_edad != "-":
            st.write(f"**Referencia etaria:** {referencia_edad}")

        st.write(f"**Interpretación clínica:** {interpretacion_clinica(clasificacion)}")

        if st.button("Guardar evaluación", key="btn_guardar_evaluacion"):
            if not paciente_nombre:
                st.warning("Seleccioná un paciente antes de guardar.")
            elif percentil is None:
                st.warning("No se pudo calcular el percentil.")
            else:
                try:
                    guardar_evaluacion(
                        paciente_nombre=paciente_nombre,
                        sexo=sexo,
                        edad=edad,
                        prueba=prueba,
                        valor_medido=valor_medido,
                        percentil=percentil,
                        clasificacion=clasificacion
                    )
                    st.success("Evaluación guardada correctamente.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al guardar: {e}")

st.divider()

# =========================================================
# HISTORIALES
# =========================================================
h1, h2 = st.columns([1, 1])

with h1:
    st.markdown("## Historial corporal")

    df_inbody = obtener_historial_inbody(paciente_id)
    if not df_inbody.empty:
        df_inbody = enriquecer_historial_corporal(
            df_inbody,
            str(ficha["sexo"]).strip().lower(),
            ficha["talla_m"]
        )
        df_inbody["fecha"] = pd.to_datetime(df_inbody["fecha"], errors="coerce")
        df_inbody = df_inbody.dropna(subset=["fecha"]).sort_values("fecha", ascending=False)

        df_inbody_mostrar = agregar_identificacion_paciente(df_inbody, ficha)
        df_inbody_mostrar["fecha"] = pd.to_datetime(df_inbody_mostrar["fecha"], errors="coerce").dt.strftime("%Y-%m-%d")

        columnas_inbody = [
            "Paciente",
            "Sexo",
            "Talla_m",
            "PacienteID",
            "fecha",
            "peso_kg",
            "imc",
            "grasa_corporal_pct",
            "masa_muscular_kg",
            "agua_corporal_pct",
            "grasa_visceral",
            "metabolismo_basal",
            "diagnostico_corporal",
            "sugerencia_corporal",
            "motivos_corporal",
            "observaciones"
        ]
        columnas_inbody = [c for c in columnas_inbody if c in df_inbody_mostrar.columns]

        st.dataframe(
            df_inbody_mostrar[columnas_inbody],
            use_container_width=True,
            hide_index=True
        )
    else:
        st.info("Sin historial corporal.")

with h2:
    st.markdown("## Historial funcional")

    df_historial = obtener_historial_paciente(paciente_nombre)

    if not df_historial.empty:
        prueba_filtro = st.selectbox(
            "Filtrar historial por prueba",
            options=["Todas", "Caminata 6 minutos", "Prensión manual", "Levantarse de la silla"],
            index=0,
            key="filtro_historial_prueba"
        )

        if prueba_filtro == "Todas":
            df_historial_filtrado = df_historial.copy()
        else:
            df_historial_filtrado = df_historial[
                df_historial["prueba"].astype(str).str.strip() == prueba_filtro
            ].copy()

        columnas_mostrar = ["id", "fecha", "prueba", "valor_medido", "percentil", "clasificacion"]
        columnas_existentes = [c for c in columnas_mostrar if c in df_historial_filtrado.columns]
        df_historial_mostrar = df_historial_filtrado[columnas_existentes].copy()

        if "fecha" in df_historial_mostrar.columns:
            df_historial_mostrar["fecha"] = pd.to_datetime(
                df_historial_mostrar["fecha"],
                errors="coerce"
            ).dt.strftime("%Y-%m-%d")

        df_historial_mostrar = df_historial_mostrar.sort_values(by="fecha", ascending=False)

        st.markdown("**Fecha | Prueba | Valor | Percentil | Clasificación | Eliminar**")

        for _, row in df_historial_mostrar.iterrows():
            c1, c2, c3, c4, c5, c6 = st.columns([1, 2, 1, 1, 1, 0.5])

            c1.write(row.get("fecha", ""))
            c2.write(row.get("prueba", ""))
            c3.write(row.get("valor_medido", ""))
            c4.write(row.get("percentil", ""))
            c5.write(row.get("clasificacion", ""))

            if c6.button("🗑", key=f"del_{row['id']}"):
                try:
                    eliminar_evaluacion(row["id"])
                    st.success("Evaluación eliminada.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al eliminar: {e}")
    else:
        st.info("Sin historial funcional.")

st.divider()

# =========================================================
# GRÁFICOS
# =========================================================
st.markdown("## Gráficos")

g1, g2 = st.columns([1, 1])

with g1:
    st.markdown("### Evolución de peso e IMC")

    if paciente_id is not None:
        df_peso = obtener_historial_peso(paciente_id)

        if not df_peso.empty:
            df_peso["fecha"] = pd.to_datetime(df_peso["fecha"], errors="coerce")
            df_peso["peso_kg"] = pd.to_numeric(df_peso["peso_kg"], errors="coerce")
            df_peso["imc"] = pd.to_numeric(df_peso["imc"], errors="coerce")
            df_peso = df_peso.dropna(subset=["fecha", "peso_kg", "imc"]).sort_values("fecha", ascending=True)

            grafico_doble = alt.Chart(df_peso).transform_fold(
                ["peso_kg", "imc"],
                as_=["variable", "valor"]
            ).mark_line(point=True).encode(
                x=alt.X("yearmonthdate(fecha):T", title="Fecha"),
                y=alt.Y("valor:Q", title="Valor"),
                color=alt.Color("variable:N", title="Serie"),
                tooltip=[
                    alt.Tooltip("yearmonthdate(fecha):T", title="Fecha"),
                    alt.Tooltip("variable:N", title="Serie"),
                    alt.Tooltip("valor:Q", title="Valor", format=".2f")
                ]
            ).properties(height=320)

            st.altair_chart(grafico_doble, use_container_width=True)
        else:
            st.info("Sin datos de peso / IMC.")

with g2:
    st.markdown("### Evolución del percentil funcional")

    if paciente_nombre:
        df_historial = obtener_historial_paciente(paciente_nombre)

        if not df_historial.empty and {"fecha", "percentil", "prueba"}.issubset(df_historial.columns):
            df_graf_base = df_historial.copy()
            df_graf_base["fecha"] = pd.to_datetime(df_graf_base["fecha"], errors="coerce").dt.date
            df_graf_base["percentil"] = pd.to_numeric(df_graf_base["percentil"], errors="coerce")
            df_graf_base["prueba"] = df_graf_base["prueba"].astype(str).str.strip()
            df_graf_base = df_graf_base.dropna(subset=["fecha", "percentil", "prueba"])

            prueba_grafico = st.selectbox(
                "Prueba para gráfico",
                options=["Caminata 6 minutos", "Prensión manual", "Levantarse de la silla"],
                key="selector_grafico_prueba"
            )

            df_prueba = df_graf_base[df_graf_base["prueba"] == prueba_grafico].copy()

            if not df_prueba.empty:
                df_prueba = (
                    df_prueba.groupby("fecha", as_index=False)["percentil"]
                    .mean()
                    .sort_values("fecha")
                )

                df_prueba["Etiqueta"] = df_prueba["percentil"].apply(lambda x: f"P{round(x, 1)}")

                linea = alt.Chart(df_prueba).mark_line(point=False).encode(
                    x=alt.X(
                        "yearmonthdate(fecha):T",
                        title="Fecha",
                        axis=alt.Axis(format="%d-%m-%Y")
                    ),
                    y=alt.Y("percentil:Q", title="Percentil")
                )

                puntos = alt.Chart(df_prueba).mark_circle(size=90).encode(
                    x=alt.X("yearmonthdate(fecha):T"),
                    y=alt.Y("percentil:Q")
                )

                etiquetas = alt.Chart(df_prueba).mark_text(
                    dy=-12,
                    fontSize=12
                ).encode(
                    x=alt.X("yearmonthdate(fecha):T"),
                    y=alt.Y("percentil:Q"),
                    text="Etiqueta:N"
                )

                grafico = (linea + puntos + etiquetas).properties(height=320)
                st.altair_chart(grafico, use_container_width=True)

                ultimo = df_prueba["percentil"].iloc[-1]
                if len(df_prueba) >= 2:
                    anterior = df_prueba["percentil"].iloc[-2]
                    diferencia = round(ultimo - anterior, 1)

                    if diferencia > 0:
                        st.success(f"↑ Mejora de {diferencia} percentiles desde la evaluación anterior")
                    elif diferencia < 0:
                        st.warning(f"↓ Disminución de {abs(diferencia)} percentiles desde la evaluación anterior")
                    else:
                        st.info("Sin cambios respecto a la evaluación anterior")
            else:
                st.info("Sin datos funcionales para esa prueba.")
        else:
            st.info("Sin historial funcional para graficar.")
